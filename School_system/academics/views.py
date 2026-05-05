import logging
import csv
import io
import json
import re
import secrets
import string
import os
from decimal import Decimal

from django.db.models import Avg, Count, Q, Prefetch
from django.utils import timezone
from rest_framework import generics, status, permissions
from rest_framework.decorators import api_view, permission_classes
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.response import Response
import requests

logger = logging.getLogger(__name__)

from email_service import (
    send_result_entered_email,
    send_announcement_email,
    send_parent_link_approved_email,
    get_parents_of_student,
    send_bulk_welcome_teacher,
    send_bulk_welcome_parent,
)
from .models import (
    Subject, Class, Student, Teacher, Parent, Result, 
    Timetable, Announcement, AnnouncementDismissal, Complaint, Suspension,
    ClassAttendance, SubjectAttendance, AttendancePermission, BulkImportJob, ClassSubjectAssignment
)
from .serializers import (
    SubjectSerializer, ClassSerializer, StudentSerializer, TeacherSerializer,
    ParentSerializer, ResultSerializer, TimetableSerializer, AnnouncementSerializer,
    ComplaintSerializer, SuspensionSerializer, StudentPerformanceSerializer,
    CreateResultSerializer, CreateStudentSerializer, CreateTeacherSerializer, CreateParentSerializer,
    UpdateStudentSerializer, UpdateTeacherSerializer, UpdateParentSerializer,
    TransferredStudentSerializer,
)
from .utils import MAX_PARENTS_PER_CHILD, check_rate_limit, log_school_audit
from users.models import SchoolSettings


BULK_IMPORT_PARAMETER_LIBRARY = {
    "subjects": [
        {"key": "name", "label": "Subject Name", "required": True, "type": "text", "help": "e.g. Mathematics"},
        {"key": "code", "label": "Subject Code", "required": False, "type": "text", "help": "Optional. Leave blank (or omit this column entirely) and the system will auto-generate a code from the subject name."},
        {"key": "description", "label": "Description", "required": False, "type": "text"},
        {"key": "ca_weight", "label": "CA Weight", "required": False, "type": "number", "help": "Continuous Assessment weight (0-1, e.g. 0.4)"},
        {"key": "exam_weight", "label": "Exam Weight", "required": False, "type": "number", "help": "Final Exam weight (0-1, e.g. 0.6)"},
        {"key": "is_priority", "label": "Priority Subject", "required": False, "type": "boolean", "help": "true/false — priority subjects get daily timetable periods"},
    ],
    "classes": [
        {"key": "name", "label": "Class Name", "required": True, "type": "text", "help": "Unique within the academic year, e.g. Form 1A"},
        {"key": "grade", "label": "Grade Level", "required": False, "type": "number", "help": "Required for primary/combined schools (e.g. 1–7). Auto-derived from Form number for secondary/high schools."},
        {"key": "academic_year", "label": "Academic Year", "required": False, "type": "text", "help": "Defaults to current year (e.g. 2026)"},
        {"key": "class_teacher_email", "label": "Class Teacher Email", "required": False, "type": "email", "help": "Email of an existing teacher in this school"},
        {"key": "subjects", "label": "Subject Codes", "required": False, "type": "text", "help": "Comma-separated subject codes to assign to this class (subjects must already exist). e.g. MATH,ENG,SCI"},
    ],
    "teachers": [
        {"key": "first_name", "label": "First Name", "required": True, "type": "text"},
        {"key": "last_name", "label": "Last Name", "required": True, "type": "text"},
        {"key": "email", "label": "Email", "required": False, "type": "email", "help": "Auto-generated if blank"},
        {"key": "phone", "label": "Phone", "required": False, "type": "text", "help": "Zimbabwe number (any of: 0788..., +263788..., 263788...)"},
        {"key": "gender", "label": "Gender", "required": False, "type": "enum", "enum_values": ["M", "F", "O"], "help": "M / F / O"},
        {"key": "hire_date", "label": "Hire Date", "required": False, "type": "date", "help": "Defaults to today if blank"},
        {"key": "qualification", "label": "Qualification", "required": False, "type": "text", "help": "e.g. BSc Mathematics, PGCE"},
        {"key": "subjects", "label": "Subjects Taught", "required": False, "type": "text", "help": "Comma-separated subject codes or names (must already exist)"},
        {"key": "assigned_class", "label": "Assigned Class", "required": False, "type": "text", "help": "Class name to assign as class teacher (must already exist & be unassigned)"},
    ],
    "students": [
        {"key": "first_name", "label": "First Name", "required": True, "type": "text"},
        {"key": "last_name", "label": "Last Name", "required": True, "type": "text"},
        {"key": "class", "label": "Class Name", "required": False, "type": "text", "help": "Optional when class is selected in the upload wizard. If provided, it can be used as fallback."},
        {"key": "gender", "label": "Gender", "required": False, "type": "text", "help": "Male / Female / Other"},
        {"key": "date_of_birth", "label": "Date of Birth", "required": False, "type": "date"},
        {"key": "email", "label": "Email", "required": False, "type": "email", "help": "Auto-generated if blank"},
        {"key": "phone", "label": "Phone", "required": False, "type": "text"},
        {"key": "address", "label": "Address", "required": False, "type": "text"},
        {"key": "residence_type", "label": "Residence", "required": False, "type": "enum", "enum_values": ["day", "boarding"], "help": "day / boarding (must match school's accommodation type)"},
        {"key": "admission_date", "label": "Admission Date", "required": False, "type": "date", "help": "Defaults to today if blank"},
        {"key": "emergency_contact", "label": "Emergency Contact", "required": False, "type": "text"},
    ],
    "parents": [
        {"key": "first_name", "label": "First Name", "required": True, "type": "text"},
        {"key": "last_name", "label": "Last Name", "required": True, "type": "text"},
        {"key": "phone", "label": "Phone", "required": True, "type": "text", "help": "Zimbabwe number"},
        {"key": "email", "label": "Email", "required": False, "type": "email", "help": "Auto-generated if blank"},
        {"key": "occupation", "label": "Occupation", "required": False, "type": "text"},
        {"key": "child_admission_nos", "label": "Child Student Numbers", "required": False, "type": "text", "help": "Comma-separated student numbers (e.g. STU001234,STU005678). Add multiple to link more than one child."},
    ],
    "fees": [
        {"key": "student_admission_no", "label": "Student Admission Number", "required": True, "type": "text"},
        {"key": "fee_type", "label": "Fee Type", "required": True, "type": "text", "help": "Created automatically if it doesn't exist"},
        {"key": "amount", "label": "Amount Due", "required": True, "type": "number"},
        {"key": "term", "label": "Term", "required": True, "type": "enum", "enum_values": ["term_1", "term_2", "term_3"], "help": "term_1 / term_2 / term_3 (also accepts t1, term1, etc.)"},
        {"key": "academic_year", "label": "Academic Year", "required": False, "type": "text", "help": "Defaults to current year"},
        {"key": "due_date", "label": "Due Date", "required": False, "type": "date", "help": "Defaults to today if blank"},
    ],
    "attendance": [
        {"key": "student_admission_no", "label": "Student Admission Number", "required": True, "type": "text"},
        {"key": "date", "label": "Date", "required": True, "type": "date"},
        {"key": "status", "label": "Status", "required": True, "type": "enum", "enum_values": ["present", "absent", "late", "excused"]},
        {"key": "reason", "label": "Reason / Remarks", "required": False, "type": "text"},
    ],
}

_BULK_IMPORT_ROLE_MATRIX = {
    "subjects": {"admin", "hr", "superadmin"},
    "classes": {"admin", "hr", "superadmin"},
    "teachers": {"admin", "hr", "superadmin"},
    "students": {"admin", "hr", "superadmin"},
    "parents": {"admin", "hr", "superadmin"},
    "fees": {"admin", "hr", "superadmin", "accountant"},
    "attendance": {"admin", "hr", "superadmin"},
}


def _allowed_import_types_for_role(role):
    user_role = (role or "").strip().lower()
    return [
        key for key in BULK_IMPORT_PARAMETER_LIBRARY.keys()
        if user_role in _BULK_IMPORT_ROLE_MATRIX.get(key, set())
    ]


def _can_access_import_type(role, import_type):
    return (role or "").strip().lower() in _BULK_IMPORT_ROLE_MATRIX.get(import_type, set())


def _delegate_bulk_import_to_go_workers(import_type, mapped_rows, user, selected_class=None, account_strategy="random", shared_password=""):
    """Run students/results/fees imports via Go workers using CSV payloads."""
    workers_base = os.environ.get("GO_WORKERS_URL", "http://workers:8081").rstrip("/")
    endpoint_map = {
        "students": "/api/v1/bulk/students",
        "results": "/api/v1/bulk/results",
        "fees": "/api/v1/bulk/fees",
    }
    if import_type not in endpoint_map:
        raise ValueError(f"Unsupported Go worker import type: {import_type}")

    if import_type == "students":
        headers = ["full_name", "email", "phone", "class_name", "date_of_birth", "gender"]
    elif import_type == "results":
        headers = ["student_number", "subject_code", "exam_type", "score", "max_score", "term", "year"]
    else:
        headers = ["student_number", "fee_type_name", "amount", "academic_year", "academic_term"]

    csv_buf = io.StringIO()
    writer = csv.DictWriter(csv_buf, fieldnames=headers)
    writer.writeheader()

    for row in mapped_rows:
        if import_type == "students":
            first_name = (row.get("first_name") or "").strip()
            last_name = (row.get("last_name") or "").strip()
            full_name = (f"{first_name} {last_name}".strip() or (row.get("full_name") or "").strip())
            class_name = (row.get("class") or row.get("class_name") or "").strip()
            if not class_name and selected_class is not None:
                class_name = selected_class.name
            writer.writerow({
                "full_name": full_name,
                "email": (row.get("email") or "").strip(),
                "phone": (row.get("phone") or "").strip(),
                "class_name": class_name,
                "date_of_birth": (row.get("date_of_birth") or "").strip(),
                "gender": (row.get("gender") or "").strip(),
            })
        elif import_type == "results":
            writer.writerow({
                "student_number": (row.get("student_admission_no") or row.get("student_number") or "").strip(),
                "subject_code": (row.get("subject_code") or "").strip(),
                "exam_type": (row.get("exam_type") or "").strip(),
                "score": str(row.get("score") or ""),
                "max_score": str(row.get("max_score") or ""),
                "term": (row.get("term") or "").strip(),
                "year": (row.get("year") or row.get("academic_year") or "").strip(),
            })
        else:
            writer.writerow({
                "student_number": (row.get("student_admission_no") or row.get("student_number") or "").strip(),
                "fee_type_name": (row.get("fee_type") or row.get("fee_type_name") or "").strip(),
                "amount": str(row.get("amount") or ""),
                "academic_year": (row.get("academic_year") or "").strip(),
                "academic_term": (row.get("term") or row.get("academic_term") or "").strip(),
            })

    files = {"file": ("bulk_import.csv", csv_buf.getvalue().encode("utf-8"), "text/csv")}
    form_data = {}
    if import_type == "students":
        form_data["account_strategy"] = (account_strategy or "random").strip().lower()
        if form_data["account_strategy"] == "shared":
            form_data["shared_password"] = shared_password or ""
    headers = {
        "X-Gateway-Auth": "true",
        "X-User-ID": str(user.id),
        "X-User-Role": str(user.role or ""),
        "X-User-School-ID": str(user.school_id or ""),
    }
    resp = requests.post(
        f"{workers_base}{endpoint_map[import_type]}",
        files=files,
        data=form_data,
        headers=headers,
        timeout=120,
    )
    try:
        payload = resp.json()
    except Exception:
        payload = {"error": resp.text or "Worker returned non-JSON response."}
    if resp.status_code >= 400:
        raise ValueError(payload.get("error") or payload.get("detail") or "Bulk worker request failed.")
    return payload


def _parse_bulk_rows_from_upload(upload):
    name = (upload.name or "").lower()
    if name.endswith(".csv"):
        decoded = upload.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(decoded))
        return [dict(r) for r in reader]

    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise ValidationError(f"Excel parsing is unavailable on server: {exc}")
        wb = load_workbook(upload, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        out = []
        for row in rows[1:]:
            item = {}
            for i, header in enumerate(headers):
                if not header:
                    continue
                val = row[i] if i < len(row) else ""
                item[header] = "" if val is None else str(val).strip()
            out.append(item)
        return out

    raise ValidationError("Unsupported file type. Upload .csv or .xlsx")


def _parse_import_date(value, date_format='YYYY-MM-DD'):
    """Parse a date string from a bulk-import row. Honours the user-selected
    date_format dropdown but always falls back to common formats so a single
    bad row doesn't fail an entire import. Returns a datetime.date or None."""
    import datetime
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    raw = str(value).strip()
    if not raw:
        return None
    raw = raw.split(' ')[0].split('T')[0]
    primary_map = {
        'DD/MM/YYYY': ['%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y'],
        'MM/DD/YYYY': ['%m/%d/%Y', '%m-%d-%Y'],
        'YYYY-MM-DD': ['%Y-%m-%d', '%Y/%m/%d'],
    }
    primary = primary_map.get(date_format, primary_map['YYYY-MM-DD'])
    fallbacks = ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%Y/%m/%d', '%d.%m.%Y']
    seen = set()
    for fmt in primary + [f for f in fallbacks if f not in primary]:
        if fmt in seen:
            continue
        seen.add(fmt)
        try:
            return datetime.datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _parse_bool(value):
    return str(value or '').strip().lower() in ('1', 'true', 'yes', 'y', 't')


def _split_csv_field(value):
    """Split a comma/semicolon/pipe-separated cell into clean tokens."""
    if value is None:
        return []
    raw = str(value).strip()
    if not raw:
        return []
    parts = re.split(r'[,;|]+', raw)
    return [p.strip() for p in parts if p.strip()]


def _normalize_term(value):
    raw = (value or "").strip().lower().replace(" ", "_")
    if raw in ("term1", "t1"):
        return "term_1"
    if raw in ("term2", "t2"):
        return "term_2"
    if raw in ("term3", "t3"):
        return "term_3"
    if raw in ("term_1", "term_2", "term_3"):
        return raw
    return "term_1"


def _as_int(value, default=0):
    try:
        return int(float(str(value).strip()))
    except Exception:
        return default


def _normalize_phone(value):
    """Normalize a Zimbabwe phone to +263XXXXXXXXX. Strips spaces, dashes, parens.
    Accepts: +263788539918, 263 788 539 918, 0788539918, 0788 539 918, etc.
    Returns the input unchanged if it doesn't match a known pattern (validation
    will catch it downstream)."""
    raw = (value or "").strip()
    if not raw:
        return ""
    digits_only = re.sub(r"[\s\-()]+", "", raw)
    if digits_only.startswith("+263") and len(digits_only) == 13:
        return digits_only
    if digits_only.startswith("263") and len(digits_only) == 12:
        return "+" + digits_only
    if digits_only.startswith("0") and len(digits_only) == 10:
        return "+263" + digits_only[1:]
    return raw


def _generate_import_email(school, first_name, last_name, tag, row_idx):
    code = (getattr(school, "code", "") or "school").lower()
    local = f"{(first_name or 'user').strip().lower()}.{(last_name or 'import').strip().lower()}.{tag}.{row_idx}.{code}"
    local = local.replace(" ", "").replace("..", ".")
    return f"{local}@import.local"


def _normalize_mapping(mapping):
    if not isinstance(mapping, dict):
        return {}
    out = {}
    for k, v in mapping.items():
        key = str(k or "").strip()
        val = str(v or "").strip()
        if not key or not val:
            continue
        out[key] = val
    return out


def _normalize_header_key(value):
    """Normalize incoming CSV/XLSX header keys to a predictable snake_case token."""
    token = re.sub(r'[^a-z0-9]+', '_', str(value or '').strip().lower()).strip('_')
    return token


_HEADER_ALIASES = {
    "firstname": "first_name",
    "lastname": "last_name",
    "surname": "last_name",
    "fullname": "full_name",
    "class_name": "class",
    "student_class": "class",
    "classname": "class_name",
    "subjectname": "subject_name",
    "subjectcode": "subject_code",
    "teacheremail": "teacher_email",
    "admission_no": "student_admission_no",
    "admission_number": "student_admission_no",
    "student_number": "student_admission_no",
    "student_admission_number": "student_admission_no",
    "child_admission_no": "child_admission_nos",
    "fee": "fee_type",
}


def _normalized_row_lookup(row):
    """Build a row lookup that accepts both raw and normalized/aliased header names."""
    lookup = {}
    for raw_key, raw_val in (row or {}).items():
        key = str(raw_key or "").strip()
        if key:
            lookup[key] = raw_val
        norm_key = _normalize_header_key(raw_key)
        if norm_key and norm_key not in lookup:
            lookup[norm_key] = raw_val
        alias_key = _HEADER_ALIASES.get(norm_key)
        if alias_key and alias_key not in lookup:
            lookup[alias_key] = raw_val
    return lookup


def _map_row_to_parameters(row, mapping):
    row_lookup = _normalized_row_lookup(row)
    if not mapping:
        # Preserve original row shape while also exposing normalized/aliased keys.
        return row_lookup
    mapped = {}
    # Accept both parameter->header and header->parameter shapes.
    for mk, mv in mapping.items():
        source_key = str(mk or "").strip()
        target_key = str(mv or "").strip()
        source_val = row_lookup.get(source_key, row_lookup.get(_normalize_header_key(source_key), ""))
        target_val = row_lookup.get(target_key, row_lookup.get(_normalize_header_key(target_key), ""))

        if source_key in row_lookup and target_key not in row_lookup:
            mapped[target_key] = source_val
        else:
            mapped[source_key] = target_val
    return mapped


def _get_request_ip(request):
    xff = request.META.get('HTTP_X_FORWARDED_FOR')
    if xff:
        return xff.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR')


def _should_apply_row_for_strategy(strategy, exists, row_idx, label, errors):
    normalized = (strategy or "skip").strip().lower()
    if normalized not in ("skip", "update", "error"):
        normalized = "skip"
    if not exists:
        return True
    if normalized == "update":
        return True
    if normalized == "skip":
        return False
    errors.append({"row": row_idx, "error": f"Duplicate found for {label}."})
    return False


# Subject Views
class SubjectListCreateView(generics.ListCreateAPIView):
    queryset = Subject.objects.all()
    serializer_class = SubjectSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.school:
            return (
                Subject.objects
                .filter(school=user.school)
                .prefetch_related('teachers__user')
                .order_by('name', 'id')
            )
        return Subject.objects.none()

    def perform_create(self, serializer):
        serializer.save(school=self.request.user.school)


class SubjectDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Subject.objects.all()
    serializer_class = SubjectSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.school:
            return (
                Subject.objects
                .filter(school=user.school)
                .prefetch_related('teachers__user')
                .order_by('name', 'id')
            )
        return Subject.objects.none()


# Class Views
class ClassListCreateView(generics.ListCreateAPIView):
    queryset = Class.objects.all()
    serializer_class = ClassSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.school:
            queryset = Class.objects.filter(school=user.school).select_related('class_teacher').annotate(
                _student_count=Count('students', distinct=True)
            )
        else:
            queryset = Class.objects.none()
        level_type = self.request.query_params.get('level', None)
        if level_type == 'primary':
            queryset = queryset.filter(grade_level__lte=7)
        elif level_type == 'secondary':
            queryset = queryset.filter(grade_level__gt=7)
        return queryset.order_by('grade_level', 'name', 'id')

    def perform_create(self, serializer):
        serializer.save(school=self.request.user.school)


class ClassDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Class.objects.all()
    serializer_class = ClassSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.school:
            return (
                Class.objects
                .filter(school=user.school)
                .select_related('class_teacher')
                .annotate(_student_count=Count('students', distinct=True))
                .order_by('grade_level', 'name', 'id')
            )
        return Class.objects.none()


# Student Views
class StudentListView(generics.ListCreateAPIView):
    queryset = Student.objects.all()
    permission_classes = [permissions.IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return CreateStudentSerializer
        return StudentSerializer

    def get_queryset(self):
        user = self.request.user
        if user.school:
            queryset = Student.objects.filter(user__school=user.school).select_related(
                'user', 'student_class'
            ).prefetch_related('parents__user')
        else:
            queryset = Student.objects.none()

        search_q = self.request.query_params.get('q', '').strip()
        if search_q:
            queryset = queryset.filter(
                Q(user__student_number__icontains=search_q) |
                Q(user__first_name__icontains=search_q) |
                Q(user__last_name__icontains=search_q)
            )

        class_id = self.request.query_params.get('class', None)
        if class_id:
            queryset = queryset.filter(student_class_id=class_id)

        residence_type = (self.request.query_params.get('residence_type') or '').strip().lower()
        if residence_type:
            normalize_residence = {
                'boarder': 'boarding',
                'boarders': 'boarding',
                'boarding': 'boarding',
                'day': 'day',
            }
            normalized = normalize_residence.get(residence_type)
            if normalized in ('day', 'boarding'):
                queryset = queryset.filter(residence_type=normalized)
        return queryset

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        student = serializer.save()
        payload = StudentSerializer(student, context={'request': request}).data
        limit_info = serializer.context.get('limit_exceeded_info')
        if limit_info:
            payload['limit_exceeded'] = True
            payload['limit_message'] = limit_info['message']
            payload['student_limit'] = limit_info['student_limit']
            payload['active_students'] = limit_info['active_students']
        return Response(payload, status=status.HTTP_201_CREATED)


class StudentDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Student.objects.all()
    permission_classes = [permissions.IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method in ['PUT', 'PATCH']:
            return UpdateStudentSerializer
        return StudentSerializer

    def get_queryset(self):
        user = self.request.user
        if user.school:
            return Student.objects.filter(user__school=user.school).select_related(
                'user', 'student_class'
            ).prefetch_related('parents__user')
        return Student.objects.none()

    def perform_update(self, serializer):
        if self.request.user.role not in ('admin', 'hr', 'superadmin'):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Only admin/HR can edit students.')
        serializer.save()

    def destroy(self, request, *args, **kwargs):
        if request.user.role not in ('admin', 'hr', 'superadmin'):
            return Response({'error': 'Only admin/HR can delete students.'}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def transfer_student(request, pk):
    if request.user.role not in ('admin', 'hr', 'superadmin'):
        return Response({'error': 'Only admin/HR can transfer students.'}, status=status.HTTP_403_FORBIDDEN)
    try:
        student = Student.objects.get(pk=pk, user__school=request.user.school)
    except Student.DoesNotExist:
        return Response({'error': 'Student not found.'}, status=status.HTTP_404_NOT_FOUND)

    note = request.data.get('transfer_note', '').strip()
    student.is_transferred = True
    student.transferred_at = timezone.now()
    student.transferred_by = request.user
    student.transfer_note = note
    student.save(update_fields=['is_transferred', 'transferred_at', 'transferred_by', 'transfer_note'])
    return Response({'message': 'Student transferred successfully.'}, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def past_students_search(request):
    if request.user.role not in ('admin', 'hr', 'superadmin'):
        return Response({'error': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
    q = request.query_params.get('q', '').strip()
    if not q:
        return Response({'error': 'A search term is required.'}, status=status.HTTP_400_BAD_REQUEST)
    queryset = Student.objects.transferred_only().filter(
        user__school=request.user.school
    ).filter(
        Q(user__student_number__icontains=q) |
        Q(user__first_name__icontains=q) |
        Q(user__last_name__icontains=q)
    ).select_related('user', 'student_class', 'transferred_by')
    serializer = TransferredStudentSerializer(queryset, many=True)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def student_performance_view(request, student_id):
    try:
        student = Student.objects.get(id=student_id)
        
        # Verify student belongs to same school as requesting user (tenant isolation)
        if request.user.school and student.user.school != request.user.school:
            return Response({'error': 'Permission denied'}, status=status.HTTP_403_FORBIDDEN)
        
        # Check permissions - students can only view their own, parents can view their children's
        if request.user.role == 'student' and request.user.student.id != student_id:
            return Response({'error': 'Permission denied'}, status=status.HTTP_403_FORBIDDEN)
        elif request.user.role == 'parent':
            if not request.user.parent.children.filter(id=student_id).exists():
                return Response({'error': 'Permission denied'}, status=status.HTTP_403_FORBIDDEN)
        
        academic_year = request.query_params.get('academic_year')
        academic_term = request.query_params.get('academic_term')
        
        results = Result.objects.filter(student=student)
        if academic_year:
            results = results.filter(academic_year=academic_year)
        if academic_term:
            results = results.filter(academic_term=academic_term)
        
        if not results.exists():
            return Response({'message': 'No results found for this student'})

        # Composite per-subject percentages via plan weights (with equal-weight
        # fallback where no plan is attached). Then take the mean across subjects.
        from .grading_calc import compute_from_queryset
        per_subject = compute_from_queryset(results.select_related('assessment_plan'))
        average_percentage = (
            sum(per_subject.values()) / len(per_subject) if per_subject else 0
        )

        # Determine overall grade — Zimbabwe grading system
        from .grading import percentage_to_grade
        grade_info = percentage_to_grade(average_percentage)

        performance_data = {
            'student_id': student.id,
            'student_name': student.user.full_name,
            'student_number': student.user.student_number,
            'class_name': student.student_class.name,
            'academic_year': academic_year or 'All Years',
            'academic_term': academic_term or 'All Terms',
            'total_subjects': results.values('subject').distinct().count(),
            'average_score': round(average_percentage, 2),
            'overall_grade': grade_info['grade'],
            'grade_description': grade_info['description'],
            'passed': grade_info['passed'],
            'at_risk': grade_info['at_risk'],
            'results': ResultSerializer(results, many=True).data
        }
        
        return Response(performance_data)
        
    except Student.DoesNotExist:
        return Response({'error': 'Student not found'}, status=status.HTTP_404_NOT_FOUND)


# Teacher Views
class TeacherListView(generics.ListCreateAPIView):
    queryset = Teacher.objects.all().order_by('user__first_name', 'user__last_name', 'id')
    permission_classes = [permissions.IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return CreateTeacherSerializer
        return TeacherSerializer

    def get_queryset(self):
        user = self.request.user
        if user.school:
            return (
                Teacher.objects
                .filter(user__school=user.school)
                .select_related('user')
                .prefetch_related('subjects_taught', 'teaching_classes')
                .order_by('user__first_name', 'user__last_name', 'id')
            )
        return Teacher.objects.none()


class TeacherDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Teacher.objects.all().order_by('user__first_name', 'user__last_name', 'id')
    permission_classes = [permissions.IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method in ['PUT', 'PATCH']:
            return UpdateTeacherSerializer
        return TeacherSerializer

    def get_queryset(self):
        user = self.request.user
        if user.school:
            return (
                Teacher.objects
                .filter(user__school=user.school)
                .select_related('user')
                .prefetch_related('subjects_taught', 'teaching_classes')
                .order_by('user__first_name', 'user__last_name', 'id')
            )
        return Teacher.objects.none()

    def perform_update(self, serializer):
        if self.request.user.role not in ('admin', 'hr', 'superadmin'):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Only admin/HR can edit teachers.')
        serializer.save()

    def destroy(self, request, *args, **kwargs):
        if request.user.role not in ('admin', 'hr', 'superadmin'):
            return Response({'error': 'Only admin/HR can delete teachers.'}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)


# Parent Views
class ParentListView(generics.ListCreateAPIView):
    queryset = Parent.objects.all()
    permission_classes = [permissions.IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return CreateParentSerializer
        return ParentSerializer

    def get_queryset(self):
        user = self.request.user
        if user.school:
            return Parent.objects.filter(
                Q(user__school=user.school) |
                Q(schools=user.school) |
                Q(children__user__school=user.school)
            ).distinct().select_related('user').prefetch_related('children__user', 'children__student_class')
        return Parent.objects.none()


class ParentDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Parent.objects.all()
    permission_classes = [permissions.IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method in ['PUT', 'PATCH']:
            return UpdateParentSerializer
        return ParentSerializer

    def get_queryset(self):
        user = self.request.user
        if user.school:
            return Parent.objects.filter(
                Q(user__school=user.school) |
                Q(schools=user.school) |
                Q(children__user__school=user.school)
            ).distinct().select_related('user').prefetch_related('children__user', 'children__student_class')
        return Parent.objects.none()

    def perform_update(self, serializer):
        if self.request.user.role not in ('admin', 'hr', 'superadmin'):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Only admin/HR can edit parents.')
        serializer.save()

    def destroy(self, request, *args, **kwargs):
        if request.user.role not in ('admin', 'hr', 'superadmin'):
            return Response({'error': 'Only admin/HR can delete parents.'}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)


# Result Views
class ResultListCreateView(generics.ListCreateAPIView):
    queryset = Result.objects.all()
    permission_classes = [permissions.IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return CreateResultSerializer
        return ResultSerializer

    def get_queryset(self):
        user = self.request.user
        if user.school:
            queryset = Result.objects.filter(student__user__school=user.school).select_related(
                'student__user', 'subject', 'teacher__user'
            )
        else:
            queryset = Result.objects.none()

        # Filter by teacher if teacher is making request
        if self.request.user.role == 'teacher':
            queryset = queryset.filter(teacher__user=self.request.user)

        # Filter by student if student/parent is making request
        if self.request.user.role == 'student':
            queryset = queryset.filter(student__user=self.request.user)
        elif self.request.user.role == 'parent':
            children_ids = self.request.user.parent.children.values_list('id', flat=True)
            queryset = queryset.filter(student_id__in=children_ids)

        # Additional filters
        student_id = self.request.query_params.get('student')
        subject_id = self.request.query_params.get('subject')
        academic_year = self.request.query_params.get('academic_year')
        academic_term = self.request.query_params.get('academic_term')

        if student_id:
            queryset = queryset.filter(student_id=student_id)
        if subject_id:
            queryset = queryset.filter(subject_id=subject_id)
        if academic_year:
            queryset = queryset.filter(academic_year=academic_year)
        if academic_term:
            queryset = queryset.filter(academic_term=academic_term)

        return queryset.order_by('-date_recorded')

    def perform_create(self, serializer):
        result = serializer.save()
        # Notify parents that a result has been posted for their child
        try:
            student = result.student
            school_name = student.user.school.name if student.user.school else "Your School"
            class_name = student.student_class.name if student.student_class else "N/A"
            student_name = f"{student.user.first_name} {student.user.last_name}".strip()
            teacher_name = ""
            if result.teacher and result.teacher.user:
                t = result.teacher.user
                teacher_name = f"{t.first_name} {t.last_name}".strip() or t.email
            for p in get_parents_of_student(student):
                send_result_entered_email(
                    parent_email=p['email'],
                    parent_name=p['name'],
                    school_name=school_name,
                    student_name=student_name,
                    class_name=class_name,
                    subject_name=result.subject.name if result.subject else "N/A",
                    exam_type=result.exam_type or "test",
                    score=str(result.score),
                    max_score=str(result.max_score),
                    academic_term=result.academic_term or "",
                    academic_year=result.academic_year or "",
                    teacher_name=teacher_name,
                )
        except Exception as exc:
            logger.error("Result email notification failed: %s", exc)


class ResultDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Result.objects.all()
    serializer_class = ResultSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.school:
            queryset = Result.objects.filter(student__user__school=user.school).select_related(
                'student__user', 'subject', 'teacher__user'
            )
        else:
            queryset = Result.objects.none()
        if user.role == 'teacher':
            queryset = queryset.filter(teacher__user=user)
        return queryset


# Timetable Views
class TimetableListView(generics.ListAPIView):
    queryset = Timetable.objects.all()
    serializer_class = TimetableSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.school:
            queryset = Timetable.objects.filter(class_assigned__school=user.school).select_related(
                'class_assigned', 'subject', 'teacher__user'
            )
        else:
            queryset = Timetable.objects.none()

        if user.role == 'student':
            queryset = queryset.filter(class_assigned=user.student.student_class)
        elif user.role == 'teacher':
            queryset = queryset.filter(teacher__user=user)
        elif user.role == 'parent':
            children_classes = user.parent.children.values_list('student_class', flat=True)
            queryset = queryset.filter(class_assigned_id__in=children_classes)

        class_id = self.request.query_params.get('class')
        day = self.request.query_params.get('day')

        if class_id:
            queryset = queryset.filter(class_assigned_id=class_id)
        if day:
            queryset = queryset.filter(day_of_week=day)

        return queryset.order_by('day_of_week', 'start_time')


# Announcement Views
def _announcement_feed_queryset_for_user(user, include_dismissed=False):
    if user.school:
        queryset = Announcement.objects.filter(
            is_active=True, author__school=user.school
        ).filter(
            Q(expires_at__isnull=True) | Q(expires_at__gt=timezone.now())
        ).select_related('author', 'target_class')
    else:
        return Announcement.objects.none()

    user_role = user.role

    if user_role not in ('admin', 'hr', 'superadmin'):
        audience_aliases = {
            user_role,
            f"{user_role}s",
        }
        audience_filter = Q(target_audience='all') | Q(target_audiences__contains=['all'])
        for audience in audience_aliases:
            audience_filter |= Q(target_audience=audience) | Q(target_audiences__contains=[audience])
        queryset = queryset.filter(audience_filter)

    if user_role == 'student':
        try:
            user_class_id = user.student.student_class_id
            queryset = queryset.filter(Q(target_class__isnull=True) | Q(target_class_id=user_class_id))
        except Exception:
            queryset = queryset.filter(target_class__isnull=True)
    elif user_role == 'parent':
        from .models import ParentChildLink
        child_class_ids = list(
            ParentChildLink.objects.filter(parent=user.parent, is_confirmed=True)
            .values_list('student__student_class_id', flat=True)
        )
        if child_class_ids:
            queryset = queryset.filter(Q(target_class__isnull=True) | Q(target_class_id__in=child_class_ids))
        else:
            queryset = queryset.filter(target_class__isnull=True)
    elif user_role == 'teacher':
        try:
            teacher = user.teacher
            teacher_class_ids = list(
                set(Class.objects.filter(class_teacher=user).values_list('id', flat=True)) |
                set(Timetable.objects.filter(teacher=teacher).values_list('class_assigned_id', flat=True).distinct())
            )
            if teacher_class_ids:
                queryset = queryset.filter(Q(target_class__isnull=True) | Q(target_class_id__in=teacher_class_ids))
            else:
                queryset = queryset.filter(target_class__isnull=True)
        except Exception:
            queryset = queryset.filter(target_class__isnull=True)

    if not include_dismissed:
        queryset = queryset.exclude(
            id__in=AnnouncementDismissal.objects.filter(user=user).values_list('announcement_id', flat=True)
        )

    return queryset.order_by('-date_posted')


class AnnouncementListCreateView(generics.ListCreateAPIView):
    queryset = Announcement.objects.all()
    serializer_class = AnnouncementSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return _announcement_feed_queryset_for_user(self.request.user, include_dismissed=False)

    def perform_create(self, serializer):
        if self.request.user.role not in ('admin', 'hr'):
            raise PermissionDenied('Only admin and HR can create announcements.')
        announcement = serializer.save(author=self.request.user)
        audiences = set(announcement.target_audiences or [announcement.target_audience])
        # Notify parents if target_audience includes 'all' or 'parent'
        if not {'all', 'parent', 'parents'}.intersection(audiences):
            return
        try:
            school = self.request.user.school
            if not school:
                return
            school_name = school.name
            author_user = self.request.user
            posted_by = f"{author_user.first_name} {author_user.last_name}".strip() or author_user.email
            # Get all parents in the school whose children are confirmed-linked
            from .models import ParentChildLink
            links = ParentChildLink.objects.filter(
                is_confirmed=True,
                student__user__school=school,
            ).select_related('parent__user', 'student__user', 'student__student_class').distinct()
            # Send one email per unique parent (they may have multiple children)
            notified = set()
            for link in links:
                parent_email = link.parent.user.email
                if not parent_email or parent_email in notified:
                    continue
                notified.add(parent_email)
                parent_name = f"{link.parent.user.first_name} {link.parent.user.last_name}".strip()
                student_name = f"{link.student.user.first_name} {link.student.user.last_name}".strip()
                class_name = link.student.student_class.name if link.student.student_class else "N/A"
                send_announcement_email(
                    parent_email=parent_email,
                    parent_name=parent_name,
                    school_name=school_name,
                    student_name=student_name,
                    class_name=class_name,
                    announcement_title=announcement.title,
                    announcement_body=announcement.content,
                    posted_by=posted_by,
                )
        except Exception as exc:
            logger.error("Announcement email notification failed: %s", exc)


class AnnouncementDetailView(generics.DestroyAPIView):
    queryset = Announcement.objects.all()
    serializer_class = AnnouncementSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if not user.school:
            return Announcement.objects.none()
        return Announcement.objects.filter(author__school=user.school).select_related('author')

    def perform_destroy(self, instance):
        user = self.request.user
        if user.role in ('admin', 'hr', 'superadmin') and user.school_id == instance.author.school_id:
            instance.delete()
            return
        if instance.author_id == user.id:
            instance.delete()
            return
        raise PermissionDenied('You can only delete your own announcements.')


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def dismiss_announcement(request, pk):
    announcement = _announcement_feed_queryset_for_user(request.user, include_dismissed=True).filter(id=pk).first()
    if not announcement:
        return Response({'error': 'Announcement not found'}, status=status.HTTP_404_NOT_FOUND)
    AnnouncementDismissal.objects.get_or_create(user=request.user, announcement=announcement)
    return Response({'message': 'Announcement cleared from your page.'}, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def dismiss_all_announcements(request):
    visible_ids = list(
        _announcement_feed_queryset_for_user(request.user, include_dismissed=True).values_list('id', flat=True)
    )
    if not visible_ids:
        return Response({'dismissed': 0, 'total_visible': 0}, status=status.HTTP_200_OK)

    existing = set(
        AnnouncementDismissal.objects.filter(
            user=request.user, announcement_id__in=visible_ids
        ).values_list('announcement_id', flat=True)
    )
    new_ids = [announcement_id for announcement_id in visible_ids if announcement_id not in existing]
    AnnouncementDismissal.objects.bulk_create(
        [AnnouncementDismissal(user=request.user, announcement_id=announcement_id) for announcement_id in new_ids]
    )
    return Response({'dismissed': len(new_ids), 'total_visible': len(visible_ids)}, status=status.HTTP_200_OK)


# Complaint Views
class ComplaintListCreateView(generics.ListCreateAPIView):
    queryset = Complaint.objects.all()
    serializer_class = ComplaintSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.school:
            queryset = Complaint.objects.filter(
                Q(school=user.school) | Q(student__user__school=user.school)
            ).distinct().select_related(
                'student__user', 'submitted_by'
            )
        else:
            queryset = Complaint.objects.none()

        if user.role in ('admin', 'hr', 'superadmin'):
            pass
        elif user.role == 'student':
            queryset = queryset.filter(student__user=user)
        elif user.role == 'parent':
            queryset = queryset.filter(submitted_by=user)
        elif user.role == 'teacher':
            queryset = queryset.filter(submitted_by=user)
        else:
            queryset = Complaint.objects.none()

        return queryset.order_by('-date_submitted')

    def perform_create(self, serializer):
        user = self.request.user
        if user.role not in ('admin', 'hr', 'teacher', 'parent', 'superadmin'):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Only admin, HR, teachers, and parents can create complaints.')

        student = serializer.validated_data.get('student')
        if student:
            if user.school and student.user.school_id != user.school_id:
                from rest_framework.exceptions import PermissionDenied
                raise PermissionDenied('Selected student is outside your school.')
            if user.role == 'parent' and not user.parent.children.filter(id=student.id).exists():
                from rest_framework.exceptions import PermissionDenied
                raise PermissionDenied('Parents can only file complaints for their linked children.')

        complaint_type = serializer.validated_data.get('complaint_type')
        if not complaint_type:
            complaint_type = {
                'parent': 'parent',
                'teacher': 'teacher',
            }.get(user.role, 'general')

        serializer.save(
            submitted_by=user,
            school=user.school,
            complaint_type=complaint_type,
        )


class ComplaintDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Complaint.objects.all()
    serializer_class = ComplaintSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if not user.school:
            return Complaint.objects.none()

        queryset = Complaint.objects.filter(
            Q(school=user.school) | Q(student__user__school=user.school)
        ).distinct()

        if user.role in ('admin', 'hr', 'superadmin'):
            return queryset
        if user.role in ('teacher', 'parent'):
            return queryset.filter(submitted_by=user)
        if user.role == 'student':
            return queryset.filter(student__user=user)
        return Complaint.objects.none()

    def update(self, request, *args, **kwargs):
        if request.user.role not in ('admin', 'hr', 'superadmin'):
            return Response({'error': 'Only admin/HR can update complaints.'}, status=status.HTTP_403_FORBIDDEN)
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if request.user.role not in ('admin', 'hr', 'superadmin'):
            return Response({'error': 'Only admin/HR can delete complaints.'}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)


# Suspension Views
class SuspensionListCreateView(generics.ListCreateAPIView):
    queryset = Suspension.objects.all()
    serializer_class = SuspensionSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.school:
            queryset = Suspension.objects.filter(student__user__school=user.school).select_related(
                'student__user', 'teacher__user'
            )
        else:
            queryset = Suspension.objects.none()

        if user.role == 'student':
            queryset = queryset.filter(student__user=user)
        elif user.role == 'parent':
            children_ids = user.parent.children.values_list('id', flat=True)
            queryset = queryset.filter(student_id__in=children_ids)
        elif user.role == 'teacher':
            queryset = queryset.filter(teacher__user=user)

        return queryset.order_by('-date_created')

    def perform_create(self, serializer):
        user = self.request.user
        if user.role not in ('admin', 'hr'):
            raise PermissionDenied('Only admin and HR can issue suspensions.')

        student = serializer.validated_data.get('student')
        if not student:
            raise ValidationError({'student': 'student is required'})
        if user.school and student.user.school_id != user.school_id:
            raise PermissionDenied('Selected student is outside your school.')

        selected_teacher = serializer.validated_data.get('teacher')
        if selected_teacher:
            if user.school and selected_teacher.user.school_id != user.school_id:
                raise PermissionDenied('Selected teacher is outside your school.')
            serializer.save(teacher=selected_teacher)
            return

        # If no teacher provided, try class teacher as sensible default.
        class_teacher_user = getattr(student.student_class, 'class_teacher', None)
        if class_teacher_user:
            class_teacher_profile = Teacher.objects.filter(
                user=class_teacher_user,
                user__school=user.school
            ).first()
            if class_teacher_profile:
                serializer.save(teacher=class_teacher_profile)
                return

        raise ValidationError({
            'teacher': (
                'A teacher must be selected to record this suspension, or the student class must have a class teacher profile.'
            )
        })


# Admin Parent-Child Link Management Views
@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def pending_parent_link_requests(request):
    """Get all pending parent-child link requests (Admin only) - filtered by school"""
    if request.user.role not in ('admin', 'hr', 'superadmin'):
        return Response({'error': 'Only admin/HR can view pending requests'}, 
                       status=status.HTTP_403_FORBIDDEN)
    
    from .models import ParentChildLink
    
    school = request.user.school
    if not school:
        return Response([])
    
    pending_links = ParentChildLink.objects.filter(
        is_confirmed=False,
        student__user__school=school
    ).select_related('parent__user', 'student__user', 'student__student_class')
    
    data = []
    for link in pending_links:
        data.append({
            'id': link.id,
            'parent_id': link.parent.id,
            'parent_name': f"{link.parent.user.first_name} {link.parent.user.last_name}",
            'parent_email': link.parent.user.email,
            'student_id': link.student.id,
            'student_name': f"{link.student.user.first_name} {link.student.user.last_name}",
            'student_number': link.student.user.student_number or '',
            'class_name': link.student.student_class.name if link.student.student_class else 'Not Assigned',
            'created_at': link.linked_date,
        })
    
    return Response(data)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def approve_parent_link_request(request, link_id):
    """Approve a parent-child link request (Admin only) - filtered by school"""
    if request.user.role not in ('admin', 'hr', 'superadmin'):
        return Response({'error': 'Only admin/HR can approve requests'}, 
                       status=status.HTTP_403_FORBIDDEN)
    
    from .models import ParentChildLink
    
    school = request.user.school
    if not school:
        return Response({'error': 'No school associated with user'}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        link = ParentChildLink.objects.select_related(
            'parent__user', 'student__user'
        ).get(id=link_id, is_confirmed=False, student__user__school=school)

        current_parent_count = Parent.objects.filter(children=link.student).count()
        if current_parent_count >= MAX_PARENTS_PER_CHILD:
            return Response(
                {'error': f'Cannot approve link: this student already has {MAX_PARENTS_PER_CHILD} parents linked.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        link.is_confirmed = True
        link.confirmed_date = timezone.now()
        link.save()

        # Add the child to parent's children M2M
        link.parent.children.add(link.student)

        # Add the child's school to parent's schools M2M
        # This supports parents with children at multiple schools
        child_school = link.student.user.school
        if child_school:
            link.parent.schools.add(child_school)
            if not link.parent.user.school_id:
                link.parent.user.school = child_school
                link.parent.user.save(update_fields=['school'])
        log_school_audit(
            user=request.user,
            action='APPROVE',
            model_name='ParentChildLink',
            object_id=link.id,
            object_repr=f"Approved parent {link.parent_id} -> student {link.student_id}",
            changes={'is_confirmed': True, 'student_id': link.student_id, 'parent_id': link.parent_id},
            status_code=status.HTTP_200_OK,
            ip_address=request.META.get('REMOTE_ADDR'),
        )

        parent_name = f"{link.parent.user.first_name} {link.parent.user.last_name}".strip()
        student_name = f"{link.student.user.first_name} {link.student.user.last_name}".strip()
        school_name = school.name
        class_name = link.student.student_class.name if link.student.student_class else "N/A"
        student_number = (link.student.user.student_number or "").strip()
        student_username = (link.student.user.username or "").strip()
        student_email = (link.student.user.email or "").strip()

        # Notify parent their link was approved
        try:
            if link.parent.user.email:
                send_parent_link_approved_email(
                    parent_email=link.parent.user.email,
                    parent_name=parent_name,
                    school_name=school_name,
                    student_name=student_name,
                    class_name=class_name,
                    student_number=student_number,
                    student_username=student_username,
                    student_email=student_email,
                )
        except Exception as exc:
            logger.error("Parent link approval email failed: %s", exc)

        return Response({
            'message': 'Parent-child link approved successfully',
            'parent_name': parent_name,
            'student_name': student_name,
        })
    except ParentChildLink.DoesNotExist:
        return Response({'error': 'Link request not found or already confirmed'},
                       status=status.HTTP_404_NOT_FOUND)


@api_view(['DELETE'])
@permission_classes([permissions.IsAuthenticated])
def decline_parent_link_request(request, link_id):
    """Decline/delete a parent-child link request (Admin only) - filtered by school"""
    if request.user.role not in ('admin', 'hr', 'superadmin'):
        return Response({'error': 'Only admin/HR can decline requests'}, 
                       status=status.HTTP_403_FORBIDDEN)
    
    from .models import ParentChildLink
    
    school = request.user.school
    if not school:
        return Response({'error': 'No school associated with user'}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        link = ParentChildLink.objects.get(id=link_id, is_confirmed=False, student__user__school=school)
        parent_name = f"{link.parent.user.first_name} {link.parent.user.last_name}"
        student_name = f"{link.student.user.first_name} {link.student.user.last_name}"
        link_pk = link.id
        parent_pk = link.parent_id
        student_pk = link.student_id
        link.delete()
        log_school_audit(
            user=request.user,
            action='DELETE',
            model_name='ParentChildLink',
            object_id=link_pk,
            object_repr=f"Declined parent {parent_pk} -> student {student_pk}",
            changes={'is_confirmed': False, 'student_id': student_pk, 'parent_id': parent_pk},
            status_code=status.HTTP_200_OK,
            ip_address=request.META.get('REMOTE_ADDR'),
        )
        
        return Response({
            'message': 'Parent-child link request declined',
            'parent_name': parent_name,
            'student_name': student_name,
        })
    except ParentChildLink.DoesNotExist:
        return Response({'error': 'Link request not found or already confirmed'}, 
                       status=status.HTTP_404_NOT_FOUND)

# Class Average Results View
@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def class_averages_view(request):
    """Get class averages grouped by class and subject - filtered by school"""
    from django.db.models import Avg, Count, F
    
    school = request.user.school
    
    # Get results filtered by user's school
    queryset = Result.objects.all()
    if school:
        queryset = queryset.filter(student__user__school=school)
    
    averages = queryset.values(
        'student__student_class__name',
        'student__student_class__id',
        'subject__name',
        'subject__id',
        'exam_type'
    ).annotate(
        class_name=F('student__student_class__name'),
        subject_name=F('subject__name'),
        average_score=Avg('score'),
        average_max_score=Avg('max_score'),
        student_count=Count('student', distinct=True)
    ).order_by('class_name', 'subject_name')
    
    # Calculate percentages and grades
    results = []
    for avg in averages:
        if avg['average_max_score'] and avg['average_max_score'] > 0:
            percentage = round((avg['average_score'] / avg['average_max_score']) * 100, 2)
        else:
            percentage = 0
            
        # Calculate grade — Zimbabwe grading system
        from .grading import percentage_to_grade
        grade_info = percentage_to_grade(percentage)

        results.append({
            'class_name': avg['class_name'],
            'subject_name': avg['subject_name'],
            'exam_type': avg['exam_type'],
            'average_score': round(avg['average_score'], 2),
            'average_max_score': round(avg['average_max_score'], 2),
            'percentage': percentage,
            'grade': grade_info['grade'],
            'grade_description': grade_info['description'],
            'passed': grade_info['passed'],
            'student_count': avg['student_count']
        })
    
    return Response(results)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def generate_timetable_view(request):
    """Generate timetables for all classes using CSP algorithm - filtered by school"""
    if request.user.role not in ('admin', 'hr', 'superadmin'):
        return Response({'error': 'Only admin/HR can generate timetables'}, status=status.HTTP_403_FORBIDDEN)
    
    school = request.user.school
    if not school:
        return Response({'error': 'No school associated with user'}, status=status.HTTP_400_BAD_REQUEST)
    
    academic_year = request.data.get('academic_year')
    clear_existing = request.data.get('clear_existing', True)
    
    try:
        from .timetable_generator import generate_timetable
        
        success, message, entries = generate_timetable(
            school=school,
            academic_year=academic_year,
            clear_existing=clear_existing
        )
        
        if success:
            return Response({
                'success': True,
                'message': message,
                'entries_count': len(entries),
                'timetables': TimetableSerializer(entries, many=True).data
            })
        else:
            return Response({
                'success': False,
                'message': message
            }, status=status.HTTP_400_BAD_REQUEST)
    
    except Exception as e:
        return Response({
            'success': False,
            'message': f'Error generating timetable: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def get_timetable_stats(request):
    """Get timetable statistics for admin - filtered by school"""
    if request.user.role not in ('admin', 'hr', 'superadmin'):
        return Response({'error': 'Only admin/HR can view timetable stats'}, status=status.HTTP_403_FORBIDDEN)
    
    school = request.user.school
    if not school:
        return Response({'error': 'No school associated with user'}, status=status.HTTP_400_BAD_REQUEST)
    
    total_entries = Timetable.objects.filter(class_assigned__school=school).count()
    classes_with_timetables = Timetable.objects.filter(class_assigned__school=school).values('class_assigned').distinct().count()
    total_classes = Class.objects.filter(school=school).count()
    
    return Response({
        'total_entries': total_entries,
        'classes_with_timetables': classes_with_timetables,
        'total_classes': total_classes,
        'coverage_percent': round((classes_with_timetables / total_classes * 100) if total_classes > 0 else 0, 1)
    })


# ---------------------------------------------------------------
# Report Card PDF Generation
# ---------------------------------------------------------------

def _normalize_report_year(year):
    return str(year or '').strip()


def _normalize_report_term(term):
    raw = str(term or '').strip()
    lowered = raw.lower().replace('_', ' ')
    compact = lowered.replace(' ', '')
    mapping = {
        'term1': 'Term 1',
        'term2': 'Term 2',
        'term3': 'Term 3',
        '1': 'Term 1',
        '2': 'Term 2',
        '3': 'Term 3',
    }
    return mapping.get(compact, raw)

def _student_has_approved_plan_for_term(student, school, year, term_key):
    """
    Interpret an approved payment plan as an installment record with a next due date.
    This allows access when a student is on an active admin/accounting-managed plan.
    """
    from finances.models import StudentPaymentRecord
    from finances.term_finance import resolve_terms_for_plan, normalize_term_key

    records = StudentPaymentRecord.objects.filter(
        school=school,
        student=student,
        academic_year=str(year),
        payment_type='school_fees',
        payment_status__in=['partial', 'unpaid'],
    ).exclude(next_payment_due__isnull=True)

    for record in records:
        covered = resolve_terms_for_plan(
            record.payment_plan,
            record.academic_term,
            getattr(record, 'covered_terms', []),
        )
        if not covered:
            fallback = normalize_term_key(record.academic_term)
            covered = [fallback] if fallback else []
        if term_key in covered:
            return True
    return False


def _student_is_financially_eligible_for_report(student, school, year, term):
    """For fully-paid releases: block students who owe current/previous term without approved plans."""
    from finances.term_finance import TERM_SEQUENCE, TERM_INDEX, normalize_term_key
    from finances.views import _student_term_financials

    term_key = normalize_term_key(term)
    if term_key not in TERM_INDEX:
        return True

    terms_to_check = TERM_SEQUENCE[:TERM_INDEX[term_key] + 1]
    expected_map, collected_map = _student_term_financials(
        students=[student],
        school=school,
        academic_year=str(year),
        terms=terms_to_check,
    )

    for current_term in terms_to_check:
        expected = expected_map.get((student.id, current_term), Decimal('0'))
        collected = collected_map.get((student.id, current_term), Decimal('0'))
        outstanding = expected - collected
        if outstanding > Decimal('0.01'):
            if not _student_has_approved_plan_for_term(student, school, year, current_term):
                return False
    return True

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def generate_report_card(request, student_id):
    """
    Generate a PDF report card for a student.
    Query params: ?year=2025&term=Term+1
    Allowed: admin (any student in school), student (own report only),
             parent (confirmed linked children only),
             teacher (students in classes they teach / are class teacher of).
    """
    from django.http import HttpResponse

    user = request.user
    school = user.school
    year = _normalize_report_year(request.query_params.get('year', ''))
    term = _normalize_report_term(request.query_params.get('term', ''))

    try:
        student = Student.objects.select_related('user', 'student_class').get(id=student_id, user__school=school)
    except Student.DoesNotExist:
        return Response({'error': 'Student not found.'}, status=status.HTTP_404_NOT_FOUND)

    # ── Permission checks per role ──────────────────────────────────────
    if user.role == 'student':
        try:
            if user.student.id != student.id:
                return Response({'error': 'You can only view your own report card.'}, status=status.HTTP_403_FORBIDDEN)
        except Student.DoesNotExist:
            return Response({'error': 'Student profile not found.'}, status=status.HTTP_403_FORBIDDEN)

    elif user.role == 'parent':
        from .models import ParentChildLink
        is_linked = ParentChildLink.objects.filter(
            parent=user.parent, student=student, is_confirmed=True
        ).exists()
        if not is_linked:
            return Response({'error': 'You can only view report cards for your confirmed children.'}, status=status.HTTP_403_FORBIDDEN)

    elif user.role == 'teacher':
        from .models import Timetable
        teacher = user.teacher
        is_class_teacher = Class.objects.filter(
            id=student.student_class_id, class_teacher=user
        ).exists() if student.student_class_id else False
        teaches_class = Timetable.objects.filter(
            teacher=teacher, class_assigned_id=student.student_class_id
        ).exists() if student.student_class_id else False
        if not is_class_teacher and not teaches_class:
            return Response({'error': 'You can only view report cards for students in your classes.'}, status=status.HTTP_403_FORBIDDEN)

    elif user.role == 'admin':
        pass  # admins can access any student in their school

    else:
        return Response({'error': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)

    # ── Check if reports are published (students and parents only) ─────
    if user.role in ('student', 'parent') and student.student_class_id:
        from .models import ReportCardRelease
        release = ReportCardRelease.objects.filter(
            school=school, class_obj_id=student.student_class_id,
            academic_year=year, academic_term=term,
        ).first()
        if not release:
            return Response(
                {'error': 'Report cards for this term have not been published yet. Please check back later.'},
                status=status.HTTP_403_FORBIDDEN
            )
        if release.access_scope == 'fully_paid':
            if not _student_is_financially_eligible_for_report(student, school, year, term):
                return Response(
                    {
                        'error': (
                            'Report access is currently limited to fully-paid students '
                            'or those on approved payment plans for current/previous terms.'
                        )
                    },
                    status=status.HTTP_403_FORBIDDEN
                )
        if student.id in _excluded_student_ids_for_delivery(school, student.student_class, year, term):
            return Response(
                {'error': 'Report delivery is temporarily blocked due to a data issue. Please contact the school office.'},
                status=status.HTTP_403_FORBIDDEN
            )

    # ── Build PDF ───────────────────────────────────────────────────────
    # Only include results marked for the report card.
    # Use report_term override when set, otherwise fall back to academic_term.
    from django.db.models import Case, When, F, CharField
    results = Result.objects.filter(
        student=student, academic_year=year, include_in_report=True,
    ).annotate(
        effective_term=Case(
            When(report_term='', then=F('academic_term')),
            default=F('report_term'),
            output_field=CharField(),
        )
    ).filter(effective_term=term).select_related('subject').order_by('subject__name')

    buffer = _build_report_card_pdf(student, results, school, year, term)

    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="report_card_{student.user.student_number}_{term}_{year}.pdf"'
    )
    return response


def _get_report_config(school):
    """Get the ReportCardConfig for a school, or return None for defaults."""
    from users.models import ReportCardConfig
    try:
        return ReportCardConfig.objects.get(school=school)
    except ReportCardConfig.DoesNotExist:
        return None


def _cfg(cfg, attr, default):
    """Safe getattr for config or default."""
    return getattr(cfg, attr, default) if cfg else default


def _font_name(family, bold=False, italic=False):
    """Map font_family config → reportlab built-in font."""
    if family == 'sans':
        if bold and italic:
            return 'Helvetica-BoldOblique'
        if bold:
            return 'Helvetica-Bold'
        if italic:
            return 'Helvetica-Oblique'
        return 'Helvetica'
    elif family == 'elegant':
        italic = True  # elegant is italic-leaning

    # serif/elegant families map to Times built-ins
    if bold and italic:
        return 'Times-BoldItalic'
    if bold:
        return 'Times-Bold'
    if italic:
        return 'Times-Italic'
    return 'Times-Roman'


def _font_scale(scale):
    return {'compact': 0.88, 'normal': 1.0, 'large': 1.12}.get(scale, 1.0)


def _compute_class_position(student, year, term):
    """Return (rank, class_size) for this student in their class for the term."""
    from django.db.models import Sum, F, FloatField, ExpressionWrapper
    if not student.student_class_id:
        return None, None
    class_students = Student.objects.filter(
        student_class_id=student.student_class_id, user__is_active=True,
    ).values_list('id', flat=True)
    totals = {}
    for r in Result.objects.filter(
        student_id__in=class_students, academic_year=year,
        academic_term=term, include_in_report=True, max_score__gt=0,
    ).values('student_id', 'score', 'max_score'):
        pct = (r['score'] / r['max_score']) * 100 if r['max_score'] else 0
        totals.setdefault(r['student_id'], []).append(pct)
    averages = [(sid, sum(v) / len(v)) for sid, v in totals.items() if v]
    if not averages:
        return None, None
    averages.sort(key=lambda x: x[1], reverse=True)
    for i, (sid, _) in enumerate(averages, start=1):
        if sid == student.id:
            return i, len(averages)
    return None, len(averages)


def _previous_term(term):
    return {'Term 2': 'Term 1', 'Term 3': 'Term 2'}.get(term)


def _previous_term_averages(student, year, prev_term):
    """Return {subject_name: pct} for the student's previous term."""
    if not prev_term:
        return {}
    out = {}
    for r in Result.objects.filter(
        student=student, academic_year=year, academic_term=prev_term,
        include_in_report=True,
    ).select_related('subject'):
        out.setdefault(r.subject.name, []).append(
            (r.score / r.max_score * 100) if r.max_score else 0.0
        )
    return {k: round(sum(v) / len(v), 1) for k, v in out.items() if v}


def _class_subject_stats(student, year, term):
    """Return {subject_name: (avg, high)} across the class for each subject."""
    if not student.student_class_id:
        return {}
    out = {}
    rows = Result.objects.filter(
        student__student_class_id=student.student_class_id,
        academic_year=year, academic_term=term, include_in_report=True, max_score__gt=0,
    ).select_related('subject').values('subject__name', 'score', 'max_score', 'student_id')
    by_subj = {}
    for r in rows:
        pct = (r['score'] / r['max_score']) * 100
        by_subj.setdefault(r['subject__name'], {}).setdefault(r['student_id'], []).append(pct)
    for subj, per_student in by_subj.items():
        student_avgs = [sum(v) / len(v) for v in per_student.values() if v]
        if student_avgs:
            out[subj] = (round(sum(student_avgs) / len(student_avgs), 1),
                         round(max(student_avgs), 1))
    return out


def _build_report_card_pdf(student, results, school, year, term):
    """Build a single student report card PDF and return a BytesIO buffer (seeked to 0)."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, letter, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, Frame, PageTemplate,
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from .grading import percentage_to_grade, score_to_percentage
    from .models import ReportCardApprovalRequest, SubjectTermFeedback, PromotionRecord
    from io import BytesIO
    import os

    cfg = _get_report_config(school)

    # ── Config values (with defaults) ───────────────────────────────
    primary = _cfg(cfg, 'primary_color', '#1d4ed8')
    secondary = _cfg(cfg, 'secondary_color', '#f3f4f6')
    grad_start = _cfg(cfg, 'gradient_start_color', primary)
    grad_end = _cfg(cfg, 'gradient_end_color', primary)
    header_style_kind = _cfg(cfg, 'header_style', 'solid')
    font_family = _cfg(cfg, 'font_family', 'serif')
    font_scale_k = _font_scale(_cfg(cfg, 'font_size_scale', 'normal'))
    page_size_name = _cfg(cfg, 'page_size', 'A4')
    orientation = _cfg(cfg, 'page_orientation', 'portrait')
    one_page_fit = _cfg(cfg, 'one_page_fit', False)
    if one_page_fit:
        font_scale_k *= 0.9
    show_grading_key = _cfg(cfg, 'show_grading_key', True)
    show_attendance = _cfg(cfg, 'show_attendance', True)
    show_attendance_breakdown = _cfg(cfg, 'show_attendance_breakdown', False)
    show_overall_avg = _cfg(cfg, 'show_overall_average', True)
    show_position = _cfg(cfg, 'show_position', True)
    show_class_avg = _cfg(cfg, 'show_class_average', False)
    show_prev_term = _cfg(cfg, 'show_previous_term', False)
    show_effort = _cfg(cfg, 'show_effort_grade', False)
    show_chart = _cfg(cfg, 'show_subject_chart', False)
    show_promotion = _cfg(cfg, 'show_promotion_status', False)
    show_fees_status = _cfg(cfg, 'show_fees_status', False)
    show_qr = _cfg(cfg, 'show_qr_code', False)
    grouping_on = _cfg(cfg, 'subject_grouping_enabled', False)
    show_grade_remark = _cfg(cfg, 'show_grade_remark', True)
    show_exam_types = _cfg(cfg, 'show_exam_types', True)
    highlight_pf = _cfg(cfg, 'highlight_pass_fail', False)
    principal_name = _cfg(cfg, 'principal_name', '')
    principal_title = _cfg(cfg, 'principal_title', 'Head of School')
    show_class_teacher = _cfg(cfg, 'show_class_teacher', True)
    teacher_comment = _cfg(
        cfg,
        'teacher_comments_default',
        'when the teacher adds report feedback thats what must be there'
    )
    principal_comment = _cfg(cfg, 'principal_comments_default', '')

    if student.student_class_id:
        approval = ReportCardApprovalRequest.objects.filter(
            school=school,
            class_obj_id=student.student_class_id,
            academic_year=year,
            academic_term=term,
        ).order_by('-submitted_at').first()
        if approval and approval.teacher_comment:
            teacher_comment = approval.teacher_comment

    show_next_term = _cfg(cfg, 'show_next_term_dates', True)
    footer_text = _cfg(cfg, 'custom_footer_text', '')
    watermark = _cfg(cfg, 'watermark_text', '')
    border_style = _cfg(cfg, 'border_style', 'simple')
    show_conduct = _cfg(cfg, 'show_conduct_section', False)
    show_activities = _cfg(cfg, 'show_activities_section', False)

    # ── Attendance ──
    attendance_qs = student.class_attendance_records.filter(date__isnull=False)
    attendance_total = attendance_qs.count()
    present_count = attendance_qs.filter(status='present').count()
    absent_count = attendance_qs.filter(status='absent').count()
    late_count = attendance_qs.filter(status='late').count()

    # ── Per-subject feedback (comments + effort) ──
    feedback_map = {
        fb.subject.name: fb for fb in SubjectTermFeedback.objects.filter(
            student=student, academic_year=year, academic_term=term,
        ).select_related('subject')
    }

    # ── Previous term data ──
    prev_term = _previous_term(term) if show_prev_term else None
    prev_averages = _previous_term_averages(student, year, prev_term) if prev_term else {}

    # ── Class stats ──
    class_stats = _class_subject_stats(student, year, term) if show_class_avg else {}

    # ── Subject groups ──
    subject_group_map = {}
    if grouping_on:
        from users.models import SubjectGroup
        for sg in SubjectGroup.objects.filter(school=school).select_related('subject'):
            subject_group_map[sg.subject.name] = sg.group_type

    # ── Page setup ──
    base_page = A4 if page_size_name == 'A4' else letter
    pagesize = landscape(base_page) if orientation == 'landscape' else base_page

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=pagesize, topMargin=1.3*cm, bottomMargin=1.3*cm,
                            leftMargin=1.3*cm, rightMargin=1.3*cm)
    styles = getSampleStyleSheet()
    elements = []

    primary_color = colors.HexColor(primary)
    secondary_color = colors.HexColor(secondary)
    grad_start_c = colors.HexColor(grad_start)
    grad_end_c = colors.HexColor(grad_end)

    base_font = _font_name(font_family)
    bold_font = _font_name(font_family, bold=True)

    # ── Page decorator: watermark + border ──
    def _page_decorator(canvas, doc):
        canvas.saveState()
        w, h = pagesize
        if watermark:
            canvas.setFont(bold_font, 48)
            canvas.setFillColor(colors.Color(0, 0, 0, alpha=0.04))
            canvas.translate(w / 2, h / 2)
            canvas.rotate(45)
            canvas.drawCentredString(0, 0, watermark)
            canvas.restoreState()
            canvas.saveState()
        if border_style == 'simple':
            canvas.setStrokeColor(primary_color)
            canvas.setLineWidth(1.5)
            canvas.rect(1 * cm, 1 * cm, w - 2 * cm, h - 2 * cm)
        elif border_style == 'decorative':
            canvas.setStrokeColor(primary_color)
            canvas.setLineWidth(2.5)
            canvas.rect(0.8 * cm, 0.8 * cm, w - 1.6 * cm, h - 1.6 * cm)
            canvas.setLineWidth(0.5)
            canvas.rect(1.1 * cm, 1.1 * cm, w - 2.2 * cm, h - 2.2 * cm)
        canvas.restoreState()

    doc.addPageTemplates([
        PageTemplate(
            id='decorated',
            frames=[Frame(1.3 * cm, 1.3 * cm, pagesize[0] - 2.6 * cm, pagesize[1] - 2.6 * cm, id='main')],
            onPage=_page_decorator,
        )
    ])

    # ── Header block ──
    header_font_size = 18 * font_scale_k
    sub_font_size = 11 * font_scale_k
    header_text_color = colors.white if header_style_kind in ('gradient', 'banner') else primary_color

    header_para_style = ParagraphStyle(
        'Header', parent=styles['Title'], fontName=bold_font,
        fontSize=header_font_size, spaceAfter=2, textColor=header_text_color,
        alignment=TA_CENTER,
    )
    sub_style = ParagraphStyle(
        'Sub', parent=styles['Normal'], fontName=base_font, fontSize=sub_font_size,
        spaceAfter=4, alignment=TA_CENTER,
        textColor=header_text_color,
    )

    logo_img = None
    if cfg and cfg.logo and hasattr(cfg.logo, 'path') and os.path.exists(cfg.logo.path):
        try:
            logo_img = Image(cfg.logo.path, width=2.2 * cm, height=2.2 * cm)
        except Exception:
            logo_img = None

    motto_para = None
    if cfg and hasattr(school, 'settings') and getattr(school.settings, 'school_motto', ''):
        motto_style = ParagraphStyle(
            'Motto', parent=styles['Normal'], fontName=_font_name(font_family, italic=True),
            fontSize=8 * font_scale_k, textColor=header_text_color,
            alignment=TA_CENTER, spaceAfter=4,
        )
        motto_para = Paragraph(f'<i>{school.settings.school_motto}</i>', motto_style)

    name_para = Paragraph(school.name, header_para_style)
    term_para = Paragraph(f'Student Report Card &mdash; {term} {year}', sub_style)

    text_parts = [[name_para]]
    if motto_para:
        text_parts.append([motto_para])
    text_parts.append([term_para])
    full_w = pagesize[0] - 2.6 * cm

    # Build the inner header table (logo + text arrangement)
    logo_position = _cfg(cfg, 'logo_position', 'center')
    if logo_img and logo_position in ('left', 'right'):
        logo_cell_w = 2.6 * cm
        text_cell_w = full_w - logo_cell_w
        text_col = Table(text_parts, colWidths=[text_cell_w])
        if logo_position == 'left':
            inner_data = [[logo_img, text_col]]
            inner_widths = [logo_cell_w, text_cell_w]
        else:
            inner_data = [[text_col, logo_img]]
            inner_widths = [text_cell_w, logo_cell_w]
        inner = Table(inner_data, colWidths=inner_widths)
        inner.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ]))
    else:
        center_parts = []
        if logo_img:
            logo_img.hAlign = 'CENTER'
            center_parts.append([logo_img])
        center_parts.extend(text_parts)
        inner = Table(center_parts, colWidths=[full_w])
        inner.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))

    # Wrap with a styled header based on header_style_kind
    banner_img = None
    if header_style_kind == 'banner' and cfg and cfg.banner_image and \
            hasattr(cfg.banner_image, 'path') and os.path.exists(cfg.banner_image.path):
        try:
            banner_img = Image(cfg.banner_image.path, width=full_w, height=2.8 * cm)
        except Exception:
            banner_img = None

    if banner_img:
        elements.append(banner_img)
        elements.append(Spacer(1, 0.15 * cm))
        # Use primary-coloured text on top of a second copy of inner (no bg)
        inner_plain_text_color = primary_color
        # Rebuild text pieces with primary colour for banners (readable outside the banner)
        header_para_style2 = ParagraphStyle('HeaderP', parent=header_para_style, textColor=inner_plain_text_color)
        sub_style2 = ParagraphStyle('SubP', parent=sub_style, textColor=colors.black)
        plain_parts = [[Paragraph(school.name, header_para_style2)]]
        if motto_para:
            motto_style2 = ParagraphStyle('MottoP', parent=motto_style, textColor=colors.grey)
            plain_parts.append([Paragraph(f'<i>{school.settings.school_motto}</i>', motto_style2)])
        plain_parts.append([Paragraph(f'Student Report Card &mdash; {term} {year}', sub_style2)])
        inner = Table(plain_parts, colWidths=[full_w])
        inner.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'CENTER')]))
        elements.append(inner)
    elif header_style_kind == 'gradient':
        # Draw a gradient band as a Drawing then overlay text — simplest: solid fill at average for now.
        # reportlab Drawing supports LinearGradient via shapes.Rect? We'll approximate with a solid mid colour strip.
        from reportlab.graphics.shapes import Rect, Drawing as GDraw
        from reportlab.graphics.shapes import String as GString
        mid_r = (grad_start_c.red + grad_end_c.red) / 2
        mid_g = (grad_start_c.green + grad_end_c.green) / 2
        mid_b = (grad_start_c.blue + grad_end_c.blue) / 2
        band = GDraw(full_w, 2.8 * cm)
        # 16 vertical strips to fake a gradient
        steps = 16
        for i in range(steps):
            frac = i / (steps - 1)
            r = grad_start_c.red + (grad_end_c.red - grad_start_c.red) * frac
            g = grad_start_c.green + (grad_end_c.green - grad_start_c.green) * frac
            b = grad_start_c.blue + (grad_end_c.blue - grad_start_c.blue) * frac
            band.add(Rect(i * full_w / steps, 0, full_w / steps + 0.5, 2.8 * cm,
                          fillColor=colors.Color(r, g, b), strokeColor=None))
        elements.append(band)
        elements.append(Spacer(1, -2.8 * cm))  # overlap text on top
        inner.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(inner)
    elif header_style_kind == 'solid' and header_text_color == colors.white:
        # (not possible since solid uses primary text colour — we keep plain inner)
        elements.append(inner)
    else:
        elements.append(inner)

    elements.append(Spacer(1, 0.3 * cm))

    # ── Student info table ──
    info_data = [
        ['Student Name:', student.user.full_name, 'Student Number:', student.user.student_number or '-'],
        ['Class:', student.student_class.name if student.student_class else '-', 'Gender:', student.gender or '-'],
    ]
    if show_attendance:
        att_val = f'{present_count}/{attendance_total} days'
        if show_attendance_breakdown:
            att_val = f'P:{present_count} A:{absent_count} L:{late_count} (of {attendance_total})'
        info_data.append(['Admission Date:', str(student.admission_date), 'Attendance:', att_val])
    if show_class_teacher and student.student_class and student.student_class.class_teacher:
        ct = student.student_class.class_teacher
        info_data.append(['Class Teacher:', ct.full_name, '', ''])

    if show_position:
        rank, size = _compute_class_position(student, year, term)
        if rank:
            suffix = 'th' if 10 <= rank % 100 <= 20 else {1: 'st', 2: 'nd', 3: 'rd'}.get(rank % 10, 'th')
            info_data.append(['Position in Class:', f'{rank}{suffix} of {size}', '', ''])

    if show_promotion:
        promo = PromotionRecord.objects.filter(student=student, academic_year=year).order_by('-date_processed').first()
        if promo:
            info_data.append(['Promotion Status:',
                              f'{promo.get_action_display()}'
                              + (f' → {promo.to_class.name}' if promo.to_class else ''),
                              '', ''])

    if show_fees_status:
        try:
            from finances.models import StudentFee
            fees = StudentFee.objects.filter(student=student, academic_year=year, academic_term=term)
            due = sum((f.amount_due for f in fees), 0)
            paid = sum((f.amount_paid for f in fees), 0)
            bal = due - paid
            currency = getattr(school.settings, 'currency', 'USD') if hasattr(school, 'settings') else 'USD'
            info_data.append(['Fees Status:',
                              f'{currency} {float(bal):.2f} outstanding' if bal > 0 else 'Fully Paid',
                              '', ''])
        except Exception:
            pass

    # Compute column widths based on page width
    col_total = pagesize[0] - 2.6 * cm
    info_table = Table(info_data, colWidths=[3 * cm, col_total / 2 - 3 * cm,
                                              3.5 * cm, col_total / 2 - 3.5 * cm])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e0e7ff')),
        ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#e0e7ff')),
        ('FONTNAME', (0, 0), (-1, -1), base_font),
        ('FONTNAME', (0, 0), (0, -1), bold_font),
        ('FONTNAME', (2, 0), (2, -1), bold_font),
        ('FONTSIZE', (0, 0), (-1, -1), 9 * font_scale_k),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('PADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.4 * cm))

    # ── Aggregate results per subject ──
    # Score/max_score totals are still shown on the report for transparency,
    # but the reported % comes from compute_subject_percentage so the
    # admin-configured AssessmentPlan weights (papers/tests/assignments +
    # per-paper weights) drive the number — not the raw ratio of sums.
    from collections import OrderedDict
    from .grading_calc import compute_subject_percentage
    subject_data = OrderedDict()
    subject_results = OrderedDict()
    subject_plan = {}
    for r in results:
        name = r.subject.name
        if name not in subject_data:
            subject_data[name] = {'score': 0, 'max_score': 0, 'pct': 0.0}
            subject_results[name] = []
        subject_data[name]['score'] += r.score
        subject_data[name]['max_score'] += r.max_score
        subject_results[name].append(r)
        if name not in subject_plan and getattr(r, 'assessment_plan_id', None):
            subject_plan[name] = r.assessment_plan
    for name, rows in subject_results.items():
        subject_data[name]['pct'] = compute_subject_percentage(rows, subject_plan.get(name))

    # Organise by group if enabled
    def _subject_list():
        if grouping_on:
            groups_order = ['core', 'language', 'elective', 'other']
            buckets = {g: [] for g in groups_order}
            ungrouped = []
            for subj_name, data in subject_data.items():
                g = subject_group_map.get(subj_name)
                if g in buckets:
                    buckets[g].append((subj_name, data))
                else:
                    ungrouped.append((subj_name, data))
            for g in groups_order:
                if buckets[g]:
                    yield (g.capitalize(), buckets[g])
            if ungrouped:
                yield ('Other', ungrouped)
        else:
            yield (None, list(subject_data.items()))

    # ── Build results heading + table(s) ──
    heading_style = ParagraphStyle('H2', parent=styles['Heading2'], fontName=bold_font,
                                   fontSize=13 * font_scale_k, textColor=primary_color)
    elements.append(Paragraph('Academic Results', heading_style))

    def _build_results_header():
        header = ['Subject', 'Score', 'Max', '%', 'Grade']
        if show_grade_remark:
            header.append('Remark')
        if show_effort:
            header.append('Effort')
        if show_class_avg:
            header.append('Class Avg')
            header.append('Top')
        if show_prev_term and prev_averages:
            header.append('Last Term')
            header.append('Trend')
        return header

    total_pct = 0.0
    subject_count = 0
    row_colors_all = []
    any_rows = False

    for group_label, items in _subject_list():
        if group_label:
            gh = ParagraphStyle('GH', parent=styles['Heading3'], fontName=bold_font,
                                fontSize=10 * font_scale_k, textColor=colors.HexColor('#374151'),
                                spaceBefore=4, spaceAfter=2)
            elements.append(Paragraph(group_label, gh))

        header = _build_results_header()
        rows = [header]
        row_colors = []

        for subj_name, data in items:
            pct = data.get('pct', score_to_percentage(data['score'], data['max_score']))
            gi = percentage_to_grade(pct)
            row = [subj_name,
                   str(round(data['score'], 1)),
                   str(round(data['max_score'], 1)),
                   f'{pct}%',
                   gi['grade']]
            if show_grade_remark:
                row.append(gi['description'])
            if show_effort:
                fb = feedback_map.get(subj_name)
                row.append(fb.effort_grade if fb and fb.effort_grade else '-')
            if show_class_avg:
                stats = class_stats.get(subj_name)
                row.append(f'{stats[0]}%' if stats else '-')
                row.append(f'{stats[1]}%' if stats else '-')
            if show_prev_term and prev_averages:
                prev = prev_averages.get(subj_name)
                if prev is None:
                    row.append('-')
                    row.append('-')
                else:
                    row.append(f'{prev}%')
                    if pct > prev + 1:
                        row.append('▲')
                    elif pct < prev - 1:
                        row.append('▼')
                    else:
                        row.append('►')
            rows.append(row)
            row_colors.append(gi['colour'])
            total_pct += pct
            subject_count += 1

        if len(rows) <= 1:
            continue
        any_rows = True

        # Column widths scale to fit page
        ncols = len(header)
        subj_col = 4 * cm
        remaining = col_total - subj_col
        rest_w = remaining / (ncols - 1)
        col_widths = [subj_col] + [rest_w] * (ncols - 1)

        result_table = Table(rows, colWidths=col_widths, repeatRows=1)
        table_style_cmds = [
            ('BACKGROUND', (0, 0), (-1, 0), primary_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), bold_font),
            ('FONTNAME', (0, 1), (-1, -1), base_font),
            ('FONTSIZE', (0, 0), (-1, -1), 8.5 * font_scale_k),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('PADDING', (0, 0), (-1, -1), 4),
        ]
        if not highlight_pf:
            table_style_cmds.append(
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, secondary_color])
            )
        result_table.setStyle(TableStyle(table_style_cmds))

        if highlight_pf:
            for i, colour_hex in enumerate(row_colors, start=1):
                c = colors.HexColor(colour_hex)
                light = colors.Color(c.red, c.green, c.blue, alpha=0.14)
                result_table.setStyle(TableStyle([('BACKGROUND', (0, i), (-1, i), light)]))

        elements.append(result_table)
        elements.append(Spacer(1, 0.3 * cm))
        row_colors_all.extend(row_colors)

    if not any_rows:
        elements.append(Paragraph('No results recorded for this term.', styles['Normal']))

    # ── Per-subject teacher comments ──
    subj_comments = [(name, fb) for name, fb in feedback_map.items() if fb.comment]
    if subj_comments:
        elements.append(Spacer(1, 0.2 * cm))
        sc_head = ParagraphStyle('SC', parent=styles['Heading3'], fontName=bold_font,
                                 fontSize=10 * font_scale_k, textColor=primary_color,
                                 spaceAfter=2)
        elements.append(Paragraph('Subject Teacher Comments', sc_head))
        for subj_name, fb in subj_comments:
            c_style = ParagraphStyle('C', parent=styles['Normal'], fontName=base_font,
                                     fontSize=8.5 * font_scale_k, leftIndent=8, spaceAfter=2)
            elements.append(Paragraph(f'<b>{subj_name}:</b> {fb.comment}', c_style))

    # ── Overall average ──
    if show_overall_avg and subject_count > 0:
        avg_pct = round(total_pct / subject_count, 1)
        avg_grade = percentage_to_grade(avg_pct)
        elements.append(Spacer(1, 0.15 * cm))
        elements.append(Paragraph(
            f'<b>Overall Average:</b> {avg_pct}% &mdash; Grade {avg_grade["grade"]} ({avg_grade["description"]})',
            styles['Normal'],
        ))

    # ── Subject score bar chart ──
    if show_chart and subject_count > 0:
        elements.append(Spacer(1, 0.4 * cm))
        chart_head = ParagraphStyle('CH', parent=styles['Heading3'], fontName=bold_font,
                                    fontSize=10 * font_scale_k, textColor=primary_color)
        elements.append(Paragraph('Subject Performance', chart_head))
        names = []
        values = []
        for subj_name, data in subject_data.items():
            names.append(subj_name[:10])
            values.append(data.get('pct', score_to_percentage(data['score'], data['max_score'])))
        d = Drawing(col_total, 4.5 * cm)
        bc = VerticalBarChart()
        bc.x = 30
        bc.y = 15
        bc.height = 90
        bc.width = col_total - 60
        bc.data = [values]
        bc.categoryAxis.categoryNames = names
        bc.categoryAxis.labels.fontSize = 7
        bc.categoryAxis.labels.angle = 30
        bc.categoryAxis.labels.dy = -6
        bc.valueAxis.valueMin = 0
        bc.valueAxis.valueMax = 100
        bc.valueAxis.valueStep = 25
        bc.bars[0].fillColor = primary_color
        d.add(bc)
        elements.append(d)

    # ── Conduct & Activities placeholders ──
    if show_conduct:
        elements.append(Spacer(1, 0.3 * cm))
        elements.append(Paragraph('Conduct &amp; Discipline', styles['Heading3']))
        elements.append(Paragraph('_' * 80, styles['Normal']))
    if show_activities:
        elements.append(Spacer(1, 0.3 * cm))
        elements.append(Paragraph('Extra-Curricular Activities', styles['Heading3']))
        elements.append(Paragraph('_' * 80, styles['Normal']))

    # ── General comments ──
    if teacher_comment or principal_comment:
        elements.append(Spacer(1, 0.3 * cm))
        c_style = ParagraphStyle('Comment', parent=styles['Normal'], fontName=base_font,
                                 fontSize=9 * font_scale_k, spaceAfter=5)
        if teacher_comment:
            elements.append(Paragraph(f"<b>Class Teacher's Comment:</b> {teacher_comment}", c_style))
        if principal_comment:
            elements.append(Paragraph(f"<b>Head of School's Comment:</b> {principal_comment}", c_style))

    # ── Next term dates ──
    if show_next_term and term != 'Term 3':
        try:
            sschool = school.settings
            next_num = {'Term 1': 2, 'Term 2': 3}.get(term)
            if next_num:
                ns = getattr(sschool, f'term_{next_num}_start', None)
                ne = getattr(sschool, f'term_{next_num}_end', None)
                if ns or ne:
                    elements.append(Spacer(1, 0.2 * cm))
                    parts = [f'<b>Next Term (Term {next_num}):</b>']
                    if ns:
                        parts.append(f'Opens {ns.strftime("%d %B %Y")}')
                    if ne:
                        parts.append(f'Closes {ne.strftime("%d %B %Y")}')
                    elements.append(Paragraph(' &mdash; '.join(parts), styles['Normal']))
        except Exception:
            pass

    # ── Grading key ──
    if show_grading_key:
        elements.append(Spacer(1, 0.3 * cm))
        elements.append(Paragraph('Grading Key', styles['Heading3']))
        key_data = [
            ['Grade', 'Description', 'Range'],
            ['A', 'Distinction', '70 - 100%'],
            ['B', 'Merit', '60 - 69%'],
            ['C', 'Credit (Pass)', '50 - 59%'],
            ['D', 'Satisfactory', '40 - 49%'],
            ['E', 'Fail', '0 - 39%'],
        ]
        key_table = Table(key_data, colWidths=[2 * cm, 3.5 * cm, 3 * cm])
        key_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#374151')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), bold_font),
            ('FONTNAME', (0, 1), (-1, -1), base_font),
            ('FONTSIZE', (0, 0), (-1, -1), 8 * font_scale_k),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('PADDING', (0, 0), (-1, -1), 3),
        ]))
        elements.append(key_table)

    # ── Signature + QR row ──
    qr_image = None
    if show_qr:
        try:
            from django.core.signing import TimestampSigner
            import qrcode
            signer = TimestampSigner(salt='report-card')
            token = signer.sign(f'v2|{school.id}|{student.id}|{year}|{term}')
            base_url = getattr(school, 'website', '') or 'https://tishanyq.co.zw'
            verify_url = f'{base_url.rstrip("/")}/api/v1/auth/reports/verify/{token}/'
            qr = qrcode.QRCode(box_size=4, border=1)
            qr.add_data(verify_url)
            qr.make(fit=True)
            qr_img_pil = qr.make_image(fill_color='black', back_color='white')
            qr_buf = BytesIO()
            qr_img_pil.save(qr_buf, format='PNG')
            qr_buf.seek(0)
            qr_image = Image(qr_buf, width=2.3 * cm, height=2.3 * cm)
        except Exception:
            qr_image = None

    if principal_name:
        elements.append(Spacer(1, 0.7 * cm))
        teacher_name_text = ''
        if show_class_teacher and student.student_class and student.student_class.class_teacher:
            teacher_name_text = student.student_class.class_teacher.full_name

        # three-column row: class teacher signature | qr (if any) | principal stamp+name
        stamp_img = None
        if cfg and cfg.stamp_image and hasattr(cfg.stamp_image, 'path') and os.path.exists(cfg.stamp_image.path):
            try:
                stamp_img = Image(cfg.stamp_image.path, width=1.8 * cm, height=1.8 * cm)
            except Exception:
                stamp_img = None

        sig_cells = [[
            Paragraph(f'<br/><br/>_________________________<br/><b>{teacher_name_text}</b><br/>Class Teacher', styles['Normal']),
            qr_image if qr_image else '',
            Paragraph(
                (f'<br/><br/>_________________________<br/><b>{principal_name}</b><br/>{principal_title}'),
                styles['Normal'],
            ),
        ]]
        third = col_total / 3
        sig_table = Table(sig_cells, colWidths=[third, third, third])
        sig_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'CENTER'),
            ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTSIZE', (0, 0), (-1, -1), 9 * font_scale_k),
        ]))
        elements.append(sig_table)
        if stamp_img:
            stamp_img.hAlign = 'RIGHT'
            elements.append(stamp_img)
    elif qr_image:
        elements.append(Spacer(1, 0.5 * cm))
        qr_image.hAlign = 'RIGHT'
        elements.append(qr_image)

    # ── Footer ──
    elements.append(Spacer(1, 0.4 * cm))
    footer_parts = [f'Generated on {timezone.now().strftime("%d %B %Y")}', school.name]
    if footer_text:
        footer_parts.append(footer_text)
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontName=base_font,
                                  fontSize=7.5 * font_scale_k, textColor=colors.grey,
                                  alignment=TA_CENTER)
    elements.append(Paragraph(' | '.join(footer_parts), footer_style))

    doc.build(elements)
    buffer.seek(0)
    return buffer


# ---------------------------------------------------------------
# Bulk CSV Import — Students & Results
# ---------------------------------------------------------------

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def bulk_import_parameter_catalog(request):
    allowed_types = _allowed_import_types_for_role(request.user.role)
    if not allowed_types:
        return Response({'error': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
    visible_library = {k: v for k, v in BULK_IMPORT_PARAMETER_LIBRARY.items() if k in allowed_types}
    return Response({
        "import_types": [{"key": key, "label": key.replace("_", " ").title()} for key in allowed_types],
        "parameter_library": visible_library,
        "date_formats": ["DD/MM/YYYY", "MM/DD/YYYY", "YYYY-MM-DD"],
        "duplicate_strategies": ["skip", "update", "error"],
    })


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def bulk_import_validate(request):
    import_type = (request.data.get('import_type') or '').strip().lower()
    if import_type not in BULK_IMPORT_PARAMETER_LIBRARY:
        return Response({'error': 'Invalid import_type.'}, status=status.HTTP_400_BAD_REQUEST)
    if not _can_access_import_type(request.user.role, import_type):
        return Response({'error': f'Permission denied for import type: {import_type}.'}, status=status.HTTP_403_FORBIDDEN)
    school = request.user.school
    if not school:
        return Response({'error': 'No school context found.'}, status=status.HTTP_400_BAD_REQUEST)
    selected_class = None
    class_id_raw = (request.data.get("class_id") or "").strip()
    if import_type == "students":
        if not class_id_raw:
            return Response({'error': 'Please select a class for student bulk import.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            selected_class = Class.objects.filter(school=school, id=int(class_id_raw)).first()
        except (TypeError, ValueError):
            selected_class = None
        if not selected_class:
            return Response({'error': 'Selected class was not found in this school.'}, status=status.HTTP_400_BAD_REQUEST)

    upload = request.FILES.get('file')
    if not upload:
        return Response({'error': 'No file uploaded.'}, status=status.HTTP_400_BAD_REQUEST)

    raw_selected = request.data.get('selected_parameters', [])
    raw_mapping = request.data.get('mapping', {})
    if isinstance(raw_selected, str):
        try:
            selected_parameters = json.loads(raw_selected)
        except Exception:
            selected_parameters = [s.strip() for s in raw_selected.split(',') if s.strip()]
    elif isinstance(raw_selected, list):
        selected_parameters = raw_selected
    else:
        selected_parameters = []

    selected_parameters = [str(p).strip() for p in selected_parameters if str(p).strip()]
    if isinstance(raw_mapping, str):
        try:
            mapping = _normalize_mapping(json.loads(raw_mapping))
        except Exception:
            mapping = {}
    else:
        mapping = _normalize_mapping(raw_mapping)
    library = BULK_IMPORT_PARAMETER_LIBRARY[import_type]
    known_keys = {f['key'] for f in library}
    required = [f['key'] for f in library if f.get('required')]
    if import_type == "students":
        required = [k for k in required if k != "class"]
    effective_keys = sorted(set(required + [p for p in selected_parameters if p in known_keys]))

    try:
        rows = _parse_bulk_rows_from_upload(upload)
    except Exception as exc:
        return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    errors = []
    mapped_rows = [_map_row_to_parameters(row, mapping) for row in rows]

    for idx, row in enumerate(mapped_rows, start=2):
        row_errors = []
        for field_key in required:
            val = row.get(field_key, '')
            if val is None or str(val).strip() == '':
                row_errors.append(f"Missing required field: {field_key}")
        if row_errors:
            errors.append({"row": idx, "errors": row_errors})
            continue
        if import_type == "classes":
            subjects_raw = (row.get("subjects") or "").strip()
            if subjects_raw:
                codes = [c.strip() for c in subjects_raw.split(",") if c.strip()]
                for code in codes:
                    if not Subject.objects.filter(school=school, code__iexact=code).exists():
                        row_errors.append(f"Subject code '{code}' not found — import subjects first.")
            if row_errors:
                errors.append({"row": idx, "errors": row_errors})

    sample_rows = mapped_rows[:5]
    headers = list(rows[0].keys()) if rows else []

    return Response({
        "import_type": import_type,
        "selected_class": (
            {"id": selected_class.id, "name": selected_class.name, "academic_year": selected_class.academic_year}
            if selected_class else None
        ),
        "selected_parameters": effective_keys,
        "mapping": mapping,
        "total_rows": len(rows),
        "headers": headers,
        "sample_rows": sample_rows,
        "errors": errors[:100],
        "valid": len(errors) == 0,
        "message": f"Validated {len(rows)} rows with {len(errors)} row(s) containing errors.",
    })


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def bulk_import_commit(request):
    from users.models import AuditLog
    from finances.models import FeeType, StudentFee

    import_type = (request.data.get('import_type') or '').strip().lower()
    if import_type not in BULK_IMPORT_PARAMETER_LIBRARY:
        return Response({'error': 'Invalid import_type.'}, status=status.HTTP_400_BAD_REQUEST)
    if not _can_access_import_type(request.user.role, import_type):
        return Response({'error': f'Permission denied for import type: {import_type}.'}, status=status.HTTP_403_FORBIDDEN)

    upload = request.FILES.get('file')
    if not upload:
        return Response({'error': 'No file uploaded.'}, status=status.HTTP_400_BAD_REQUEST)

    school = request.user.school
    if not school:
        return Response({'error': 'No school context found.'}, status=status.HTTP_400_BAD_REQUEST)
    selected_class = None
    class_id_raw = (request.data.get("class_id") or "").strip()
    if import_type == "students":
        if not class_id_raw:
            return Response({'error': 'Please select a class for student bulk import.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            selected_class = Class.objects.filter(school=school, id=int(class_id_raw)).first()
        except (TypeError, ValueError):
            selected_class = None
        if not selected_class:
            return Response({'error': 'Selected class was not found in this school.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        rows = _parse_bulk_rows_from_upload(upload)
    except Exception as exc:
        return Response({'error': f'Could not parse file: {exc}'}, status=status.HTTP_400_BAD_REQUEST)

    raw_mapping = request.data.get('mapping', {})
    if isinstance(raw_mapping, str):
        try:
            mapping = _normalize_mapping(json.loads(raw_mapping))
        except Exception:
            mapping = {}
    else:
        mapping = _normalize_mapping(raw_mapping)

    raw_selected = request.data.get('selected_parameters', [])
    if isinstance(raw_selected, str):
        try:
            selected_parameters = json.loads(raw_selected)
        except Exception:
            selected_parameters = [s.strip() for s in raw_selected.split(',') if s.strip()]
    elif isinstance(raw_selected, list):
        selected_parameters = raw_selected
    else:
        selected_parameters = []

    mapped_rows = [_map_row_to_parameters(row, mapping) for row in rows]
    created, updated = 0, 0
    errors = []
    changes = []
    ip_address = _get_request_ip(request)

    job = BulkImportJob.objects.create(
        school=school,
        initiated_by=request.user,
        import_type=import_type,
        file_name=getattr(upload, "name", "") or "",
        status='pending',
        selected_parameters=selected_parameters,
        mapping=mapping,
        options={
            "date_format": request.data.get("date_format", ""),
            "duplicate_strategy": request.data.get("duplicate_strategy", ""),
            "account_strategy": (request.data.get("account_strategy") or "random").strip().lower(),
            "class_id": class_id_raw,
        },
        total_rows=len(mapped_rows),
    )

    if import_type in ("students", "results", "fees"):
        try:
            account_strategy = (request.data.get("account_strategy") or "random").strip().lower()
            shared_password = (request.data.get("shared_password") or "").strip()
            if import_type == "students":
                if account_strategy not in ("random", "shared", "inactive"):
                    account_strategy = "random"
                if account_strategy == "shared" and len(shared_password) < 8:
                    raise ValueError("shared_password must be at least 8 characters when using the shared strategy.")
            worker_result = _delegate_bulk_import_to_go_workers(
                import_type=import_type,
                mapped_rows=mapped_rows,
                user=request.user,
                selected_class=selected_class,
                account_strategy=account_strategy,
                shared_password=shared_password,
            )
            created = int(worker_result.get("created", 0) or 0)
            updated = int(worker_result.get("updated", 0) or 0)
            errors = worker_result.get("errors") or []
            status_value = 'completed' if not errors else ('failed' if created == 0 and updated == 0 else 'completed')
            job.status = status_value
            job.created_count = created
            job.updated_count = updated
            job.error_count = len(errors)
            job.errors = errors[:500] if isinstance(errors, list) else []
            job.changes = []
            job.completed_at = timezone.now()
            job.save(update_fields=[
                'status', 'created_count', 'updated_count', 'error_count', 'errors', 'changes', 'completed_at'
            ])
            AuditLog.objects.create(
                user=request.user,
                school=school,
                action='CREATE',
                model_name='BulkImportJob',
                object_id=str(job.id),
                object_repr=f'{import_type} import',
                changes={
                    'import_type': import_type,
                    'created': created,
                    'updated': updated,
                    'errors': len(errors),
                    'executor': 'go-workers',
                },
                ip_address=ip_address,
                response_status=200,
            )
            return Response({
                "job_id": job.id,
                "import_type": import_type,
                "created": created,
                "updated": updated,
                "errors": errors,
                "message": worker_result.get("message") or f"Imported {created} records with {len(errors)} errors.",
            })
        except Exception as exc:
            job.status = 'failed'
            job.error_count = 1
            job.errors = [{"row": 0, "error": str(exc)}]
            job.completed_at = timezone.now()
            job.save(update_fields=['status', 'error_count', 'errors', 'completed_at'])
            return Response({'error': str(exc)}, status=status.HTTP_502_BAD_GATEWAY)

    # Account-creation strategy for student/teacher/parent imports.
    # 'random' (default) — generate a random temp password the admin must distribute manually.
    # 'shared'           — every imported user gets the same admin-supplied password.
    # 'inactive'         — user is created with an unusable password (must use forgot-password).
    account_strategy = (request.data.get("account_strategy") or "random").strip().lower()
    shared_password = (request.data.get("shared_password") or "").strip()
    if account_strategy not in ("random", "shared", "inactive"):
        account_strategy = "random"
    if account_strategy == "shared" and len(shared_password) < 8:
        return Response(
            {"error": "shared_password must be at least 8 characters when using the shared strategy."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    def _password_for_row():
        if account_strategy == "shared":
            return shared_password
        if account_strategy == "inactive":
            # Long random string the user can never know — they must reset to log in.
            return "!" + secrets.token_urlsafe(24)
        # random: a short readable temp password
        alphabet = string.ascii_letters + string.digits
        return "Tmp!" + "".join(secrets.choice(alphabet) for _ in range(8))

    date_format = (request.data.get("date_format") or "YYYY-MM-DD").strip()
    duplicate_strategy = (request.data.get("duplicate_strategy") or "skip").strip().lower()
    if duplicate_strategy not in ("skip", "update", "error"):
        duplicate_strategy = "skip"
    current_year = str(timezone.now().year)
    today = timezone.now().date()

    if import_type == "subjects":
        for i, row in enumerate(mapped_rows, start=2):
            try:
                name = (row.get("name") or "").strip()
                if not name:
                    raise ValueError("Missing subject name")
                raw_code = (row.get("code") or "").strip()
                if raw_code:
                    code = raw_code
                else:
                    base = name[:8].upper().replace(" ", "")
                    code = base
                    suffix = 1
                    while Subject.objects.filter(school=school, code=code).exists():
                        code = f"{base[:7]}{suffix}"
                        suffix += 1
                description = (row.get("description") or "").strip()
                ca_weight_raw = (row.get("ca_weight") or "").strip()
                exam_weight_raw = (row.get("exam_weight") or "").strip()
                defaults = {"name": name, "is_priority": _parse_bool(row.get("is_priority"))}
                if description:
                    defaults["description"] = description
                if ca_weight_raw:
                    try:
                        defaults["ca_weight"] = float(ca_weight_raw)
                    except ValueError:
                        raise ValueError(f"Invalid ca_weight '{ca_weight_raw}'")
                if exam_weight_raw:
                    try:
                        defaults["exam_weight"] = float(exam_weight_raw)
                    except ValueError:
                        raise ValueError(f"Invalid exam_weight '{exam_weight_raw}'")
                existing = Subject.objects.filter(school=school, code=code).first()
                if not _should_apply_row_for_strategy(duplicate_strategy, bool(existing), i, f"Subject code '{code}'", errors):
                    continue
                if existing and duplicate_strategy == "update":
                    for k, v in defaults.items():
                        setattr(existing, k, v)
                    existing.save()
                    obj, was_created = existing, False
                elif existing:
                    obj, was_created = existing, False
                else:
                    obj = Subject.objects.create(school=school, code=code, **defaults)
                    was_created = True
                if was_created:
                    created += 1
                    changes.append({"action": "create", "model": "academics.Subject", "pk": obj.pk})
                else:
                    updated += 1
                    changes.append({
                        "action": "update",
                        "model": "academics.Subject",
                        "pk": obj.pk,
                        "before": {"name": obj.name, "is_priority": obj.is_priority},
                        "after": {"name": name, "is_priority": defaults["is_priority"]},
                    })
            except Exception as exc:
                errors.append({"row": i, "error": str(exc)})

    elif import_type == "classes":
        for i, row in enumerate(mapped_rows, start=2):
            try:
                name = (row.get("name") or "").strip()
                if not name:
                    raise ValueError("Missing class name")
                grade_raw = (row.get("grade") or "").strip()
                school_type = school.school_type
                if grade_raw:
                    grade_level = _as_int(grade_raw, 0)
                    if grade_level <= 0:
                        raise ValueError("Invalid grade level (must be a positive number)")
                elif school_type in ('secondary', 'high'):
                    import re as _re
                    m = _re.match(r'form\s*(\d+)', name.lower())
                    grade_level = 7 + int(m.group(1)) if m else 8
                else:
                    raise ValueError("Grade level is required for primary/combined schools")
                academic_year = (row.get("academic_year") or current_year).strip()

                class_teacher_user = None
                teacher_email = (row.get("class_teacher_email") or "").strip().lower()
                if teacher_email:
                    from users.models import CustomUser as _CU
                    class_teacher_user = _CU.objects.filter(
                        school=school, role='teacher', email__iexact=teacher_email,
                    ).first()
                    if not class_teacher_user:
                        raise ValueError(f"Teacher with email '{teacher_email}' not found in this school")

                defaults = {"grade_level": grade_level}
                if class_teacher_user is not None:
                    defaults["class_teacher"] = class_teacher_user

                existing = Class.objects.filter(school=school, name=name, academic_year=academic_year).first()
                if not _should_apply_row_for_strategy(duplicate_strategy, bool(existing), i, f"Class '{name}' ({academic_year})", errors):
                    continue
                if existing and duplicate_strategy == "update":
                    for k, v in defaults.items():
                        setattr(existing, k, v)
                    existing.save()
                    obj, was_created = existing, False
                elif existing:
                    obj, was_created = existing, False
                else:
                    obj = Class.objects.create(school=school, name=name, academic_year=academic_year, **defaults)
                    was_created = True
                if was_created:
                    created += 1
                    changes.append({"action": "create", "model": "academics.Class", "pk": obj.pk})
                else:
                    updated += 1
                    changes.append({
                        "action": "update",
                        "model": "academics.Class",
                        "pk": obj.pk,
                        "before": {"grade_level": obj.grade_level},
                        "after": {"grade_level": grade_level},
                    })

                subjects_raw = (row.get("subjects") or "").strip()
                if subjects_raw:
                    codes = [c.strip() for c in subjects_raw.split(",") if c.strip()]
                    for code in codes:
                        subject = Subject.objects.filter(school=school, code__iexact=code).first()
                        if not subject:
                            continue
                        assignment_year = obj.academic_year or current_year
                        existing_assignment = ClassSubjectAssignment.objects.filter(
                            school=school, class_obj=obj, subject=subject, academic_year=assignment_year
                        ).first()
                        if not existing_assignment:
                            ClassSubjectAssignment.objects.create(
                                school=school,
                                class_obj=obj,
                                subject=subject,
                                academic_year=assignment_year,
                                created_by=request.user,
                            )
                            changes.append({"action": "create", "model": "academics.ClassSubjectAssignment", "pk": None})

            except Exception as exc:
                errors.append({"row": i, "error": str(exc)})

    elif import_type == "teachers":
        from .serializers import CreateTeacherSerializer
        for i, row in enumerate(mapped_rows, start=2):
            try:
                first_name = (row.get("first_name") or "").strip()
                last_name = (row.get("last_name") or "").strip()
                if not first_name or not last_name:
                    raise ValueError("Missing first_name or last_name")
                email = (row.get("email") or "").strip() or _generate_import_email(school, first_name, last_name, "teacher", i)
                hire_date = _parse_import_date(row.get("hire_date"), date_format) or today

                gender_raw = (row.get("gender") or "").strip().upper()
                gender = gender_raw if gender_raw in ("M", "F", "O", "P") else ""

                subject_tokens = _split_csv_field(row.get("subjects"))
                subject_ids = []
                for token in subject_tokens:
                    subj = Subject.objects.filter(school=school).filter(
                        Q(code__iexact=token) | Q(name__iexact=token)
                    ).first()
                    if not subj:
                        raise ValueError(f"Subject '{token}' not found in this school")
                    subject_ids.append(subj.id)

                assigned_class_id = None
                assigned_class_name = (row.get("assigned_class") or "").strip()
                if assigned_class_name:
                    assigned_class = Class.objects.filter(
                        school=school, name__iexact=assigned_class_name,
                    ).order_by('-academic_year').first()
                    if not assigned_class:
                        raise ValueError(f"Class '{assigned_class_name}' not found")
                    assigned_class_id = assigned_class.id

                raw_password = _password_for_row()
                payload = {
                    "first_name": first_name,
                    "last_name": last_name,
                    "email": email,
                    "phone_number": _normalize_phone(row.get("phone")),
                    "gender": gender,
                    "hire_date": str(hire_date),
                    "qualification": (row.get("qualification") or "").strip(),
                    "password": raw_password,
                    "is_secondary_teacher": bool(subject_ids),
                    "subject_ids": subject_ids,
                }
                if assigned_class_id:
                    payload["assigned_class_id"] = assigned_class_id

                serializer = CreateTeacherSerializer(data=payload, context={"request": request})
                if serializer.is_valid():
                    out = serializer.save()
                    created += 1
                    if isinstance(out, dict) and out.get("id"):
                        changes.append({"action": "create", "model": "academics.Teacher", "pk": out["id"]})
                    if "@import.local" not in email:
                        send_bulk_welcome_teacher(
                            email=email,
                            first_name=first_name,
                            last_name=last_name,
                            school_name=school.name,
                            password=None if account_strategy == "inactive" else raw_password,
                        )
                else:
                    raise ValueError(serializer.errors)
            except Exception as exc:
                errors.append({"row": i, "error": str(exc)})

    elif import_type == "students":
        from .serializers import CreateStudentSerializer
        school_mode = getattr(school, 'accommodation_type', 'day')
        for i, row in enumerate(mapped_rows, start=2):
            try:
                first_name = (row.get("first_name") or "").strip()
                last_name = (row.get("last_name") or "").strip()
                if not first_name and not last_name:
                    full_name = (row.get("full_name") or "").strip()
                    if full_name:
                        parts = full_name.split(" ", 1)
                        first_name = parts[0]
                        last_name = parts[1] if len(parts) > 1 else ""
                class_name = (row.get("class") or row.get("class_name") or "").strip()
                student_class = selected_class
                if student_class is None and class_name:
                    student_class = Class.objects.filter(name__iexact=class_name, school=school).first()
                if not first_name or not last_name:
                    raise ValueError("Missing first_name or last_name")
                if not student_class:
                    raise ValueError("No class selected for this upload.")

                # Residence: respect school accommodation type
                residence_raw = (row.get("residence_type") or "").strip().lower()
                if residence_raw not in ("day", "boarding"):
                    residence_raw = "boarding" if school_mode == "boarding" else "day"

                admission_date = _parse_import_date(row.get("admission_date"), date_format) or today
                dob = _parse_import_date(row.get("date_of_birth"), date_format)

                email = (row.get("email") or "").strip() or _generate_import_email(school, first_name, last_name, "student", i)
                phone_norm = _normalize_phone(row.get("phone"))

                serializer = CreateStudentSerializer(
                    data={
                        "user": {
                            "first_name": first_name,
                            "last_name": last_name,
                            "password": _password_for_row(),
                        },
                        "student_class": student_class.id,
                        "residence_type": residence_raw,
                        "admission_date": str(admission_date),
                        "student_email": email,
                        "student_contact": phone_norm,
                        "student_address": (row.get("address") or "").strip(),
                        "date_of_birth": str(dob) if dob else None,
                        "gender": (row.get("gender") or "").strip(),
                        "emergency_contact": _normalize_phone(row.get("emergency_contact")),
                    },
                    context={"request": request},
                )
                if serializer.is_valid():
                    student = serializer.save()
                    created += 1
                    if hasattr(student, "id"):
                        changes.append({"action": "create", "model": "academics.Student", "pk": student.id})
                else:
                    raise ValueError(serializer.errors)
            except Exception as exc:
                errors.append({"row": i, "error": str(exc)})

    elif import_type == "parents":
        from .serializers import CreateParentSerializer
        for i, row in enumerate(mapped_rows, start=2):
            try:
                first_name = (row.get("first_name") or "").strip()
                last_name = (row.get("last_name") or "").strip()
                phone = _normalize_phone(row.get("phone"))
                if not first_name or not last_name or not phone:
                    raise ValueError("Missing first_name, last_name, or phone")
                email = (row.get("email") or "").strip() or _generate_import_email(school, first_name, last_name, "parent", i)

                # Accept either child_admission_nos (preferred) or child_admission_no (legacy single)
                child_tokens = _split_csv_field(row.get("child_admission_nos") or row.get("child_admission_no"))
                student_ids = []
                children_info = []
                missing_children = []
                for token in child_tokens:
                    student = Student.objects.filter(
                        user__school=school, user__student_number=token,
                    ).select_related("user").first()
                    if student:
                        student_ids.append(student.id)
                        children_info.append({
                            "name": f"{student.user.first_name} {student.user.last_name}".strip(),
                            "student_number": student.user.student_number or token,
                            "email": student.user.email,
                        })
                    else:
                        missing_children.append(token)
                if missing_children:
                    raise ValueError(f"Child admission number(s) not found: {', '.join(missing_children)}")

                raw_password = _password_for_row()
                payload = {
                    "full_name": f"{first_name} {last_name}".strip(),
                    "contact_number": phone,
                    "email": email,
                    "occupation": (row.get("occupation") or "").strip(),
                    "password": raw_password,
                    "student_ids": student_ids,
                }
                serializer = CreateParentSerializer(data=payload, context={"request": request})
                if serializer.is_valid():
                    out = serializer.save()
                    created += 1
                    if hasattr(out, "id"):
                        changes.append({"action": "create", "model": "academics.Parent", "pk": out.id})
                    if "@import.local" not in email:
                        send_bulk_welcome_parent(
                            email=email,
                            first_name=first_name,
                            last_name=last_name,
                            school_name=school.name,
                            password=None if account_strategy == "inactive" else raw_password,
                            children=children_info,
                        )
                else:
                    raise ValueError(serializer.errors)
            except Exception as exc:
                errors.append({"row": i, "error": str(exc)})

    elif import_type == "fees":
        for i, row in enumerate(mapped_rows, start=2):
            try:
                admission_no = (row.get("student_admission_no") or "").strip()
                fee_type_name = (row.get("fee_type") or "").strip()
                try:
                    amount = float(row.get("amount") or 0)
                except (TypeError, ValueError):
                    raise ValueError(f"Invalid amount '{row.get('amount')}'")
                term = _normalize_term(row.get("term"))
                academic_year = (row.get("academic_year") or current_year).strip()
                if not admission_no or not fee_type_name or amount <= 0:
                    raise ValueError("Missing required fee fields (student_admission_no, fee_type, amount)")
                student = Student.objects.filter(
                    user__school=school, user__student_number=admission_no,
                ).first()
                if not student:
                    raise ValueError(f"Student '{admission_no}' not found in this school")
                fee_type, _ = FeeType.objects.get_or_create(
                    name=fee_type_name,
                    school=school,
                    defaults={"amount": amount, "academic_year": academic_year},
                )
                due_date = _parse_import_date(row.get("due_date"), date_format) or today
                existing = StudentFee.objects.filter(
                    student=student,
                    fee_type=fee_type,
                    academic_year=academic_year,
                    academic_term=term,
                ).first()
                if not _should_apply_row_for_strategy(
                    duplicate_strategy,
                    bool(existing),
                    i,
                    f"fee {admission_no}/{fee_type_name}/{academic_year}/{term}",
                    errors,
                ):
                    continue
                if existing and duplicate_strategy == "update":
                    existing.amount_due = amount
                    existing.due_date = due_date
                    existing.save()
                    fee_obj, was_created = existing, False
                elif existing:
                    fee_obj, was_created = existing, False
                else:
                    fee_obj = StudentFee.objects.create(
                        student=student,
                        fee_type=fee_type,
                        academic_year=academic_year,
                        academic_term=term,
                        amount_due=amount,
                        due_date=due_date,
                    )
                    was_created = True
                if was_created:
                    created += 1
                    changes.append({"action": "create", "model": "finances.StudentFee", "pk": fee_obj.pk})
                else:
                    updated += 1
                    changes.append({
                        "action": "update",
                        "model": "finances.StudentFee",
                        "pk": fee_obj.pk,
                        "before": {"amount_due": str(fee_obj.amount_due), "due_date": str(fee_obj.due_date)},
                        "after": {"amount_due": str(amount), "due_date": str(due_date)},
                    })
            except Exception as exc:
                errors.append({"row": i, "error": str(exc)})

    elif import_type == "attendance":
        for i, row in enumerate(mapped_rows, start=2):
            try:
                admission_no = (row.get("student_admission_no") or "").strip()
                status_val = (row.get("status") or "").strip().lower()
                if not admission_no or not status_val:
                    raise ValueError("Missing required attendance fields (student_admission_no, status)")
                if status_val not in ("present", "absent", "late", "excused"):
                    raise ValueError(f"Invalid status '{status_val}' (use: present, absent, late, excused)")
                attendance_date = _parse_import_date(row.get("date"), date_format)
                if not attendance_date:
                    raise ValueError(f"Invalid or missing date '{row.get('date')}'")
                student = Student.objects.select_related("student_class").filter(
                    user__school=school, user__student_number=admission_no,
                ).first()
                if not student:
                    raise ValueError(f"Student '{admission_no}' not found in this school")
                existing = ClassAttendance.objects.filter(student=student, date=attendance_date).first()
                if not _should_apply_row_for_strategy(
                    duplicate_strategy,
                    bool(existing),
                    i,
                    f"attendance {admission_no} on {attendance_date}",
                    errors,
                ):
                    continue
                if existing and duplicate_strategy == "update":
                    existing.class_assigned = student.student_class
                    existing.status = status_val
                    existing.remarks = (row.get("reason") or "").strip()
                    existing.recorded_by = request.user
                    existing.save()
                    att_obj, was_created = existing, False
                elif existing:
                    att_obj, was_created = existing, False
                else:
                    att_obj = ClassAttendance.objects.create(
                        student=student,
                        date=attendance_date,
                        class_assigned=student.student_class,
                        status=status_val,
                        remarks=(row.get("reason") or "").strip(),
                        recorded_by=request.user,
                    )
                    was_created = True
                if was_created:
                    created += 1
                    changes.append({"action": "create", "model": "academics.ClassAttendance", "pk": att_obj.pk})
                else:
                    updated += 1
                    changes.append({
                        "action": "update",
                        "model": "academics.ClassAttendance",
                        "pk": att_obj.pk,
                        "before": {"status": att_obj.status, "remarks": att_obj.remarks},
                        "after": {"status": status_val, "remarks": (row.get("reason") or "").strip()},
                    })
            except Exception as exc:
                errors.append({"row": i, "error": str(exc)})

    else:
        return Response({'error': f"Unknown import type: {import_type}"}, status=status.HTTP_400_BAD_REQUEST)

    status_value = 'completed' if not errors else ('failed' if created == 0 and updated == 0 else 'completed')
    job.status = status_value
    job.created_count = created
    job.updated_count = updated
    job.error_count = len(errors)
    job.errors = errors[:500]
    job.changes = changes
    job.completed_at = timezone.now()
    job.save(update_fields=[
        'status', 'created_count', 'updated_count', 'error_count', 'errors', 'changes', 'completed_at'
    ])

    AuditLog.objects.create(
        user=request.user,
        school=school,
        action='CREATE',
        model_name='BulkImportJob',
        object_id=str(job.id),
        object_repr=f'{import_type} import',
        changes={
            'import_type': import_type,
            'created': created,
            'updated': updated,
            'errors': len(errors),
        },
        ip_address=ip_address,
        response_status=200,
    )

    return Response({
        "job_id": job.id,
        "import_type": import_type,
        "created": created,
        "updated": updated,
        "errors": errors,
        "message": f"Imported {created} records (updated {updated}) with {len(errors)} errors."
    })


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def bulk_import_history(request):
    allowed_types = _allowed_import_types_for_role(request.user.role)
    if not allowed_types:
        return Response({'error': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
    school = request.user.school
    if not school:
        return Response([], status=status.HTTP_200_OK)
    jobs = BulkImportJob.objects.filter(school=school, import_type__in=allowed_types).order_by('-created_at')[:100]
    data = [{
        "id": j.id,
        "import_type": j.import_type,
        "file_name": j.file_name,
        "status": j.status,
        "total_rows": j.total_rows,
        "created_count": j.created_count,
        "updated_count": j.updated_count,
        "error_count": j.error_count,
        "created_at": j.created_at,
        "completed_at": j.completed_at,
        "rolled_back_at": j.rolled_back_at,
    } for j in jobs]
    return Response(data)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def bulk_import_rollback(request, job_id):
    from django.apps import apps
    from users.models import AuditLog

    allowed_types = _allowed_import_types_for_role(request.user.role)
    if not allowed_types:
        return Response({'error': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
    school = request.user.school
    job = BulkImportJob.objects.filter(id=job_id, school=school).first()
    if not job:
        return Response({'error': 'Import job not found.'}, status=status.HTTP_404_NOT_FOUND)
    if job.import_type not in allowed_types:
        return Response({'error': f'Permission denied for import type: {job.import_type}.'}, status=status.HTTP_403_FORBIDDEN)
    if job.status == 'rolled_back':
        return Response({'error': 'This import has already been rolled back.'}, status=status.HTTP_400_BAD_REQUEST)

    rolled_back = 0
    rollback_errors = []
    for change in reversed(job.changes or []):
        try:
            model_label = change.get('model')
            app_label, model_name = model_label.split('.', 1)
            model_cls = apps.get_model(app_label, model_name)
            pk = change.get('pk')
            action = change.get('action')
            if action == 'create':
                model_cls.objects.filter(pk=pk).delete()
                rolled_back += 1
            elif action == 'update':
                instance = model_cls.objects.filter(pk=pk).first()
                if instance:
                    before = change.get('before') or {}
                    for key, value in before.items():
                        setattr(instance, key, value)
                    instance.save()
                    rolled_back += 1
        except Exception as exc:
            rollback_errors.append(str(exc))

    job.status = 'rolled_back'
    job.rolled_back_at = timezone.now()
    job.rollback_notes = f"Rolled back items: {rolled_back}; errors: {len(rollback_errors)}"
    job.save(update_fields=['status', 'rolled_back_at', 'rollback_notes'])

    AuditLog.objects.create(
        user=request.user,
        school=school,
        action='DELETE',
        model_name='BulkImportJob',
        object_id=str(job.id),
        object_repr=f'{job.import_type} rollback',
        changes={'rolled_back': rolled_back, 'errors': rollback_errors[:20]},
        ip_address=_get_request_ip(request),
        response_status=200,
    )

    return Response({
        "job_id": job.id,
        "rolled_back": rolled_back,
        "errors": rollback_errors,
        "message": f"Rollback finished. Rolled back {rolled_back} change(s).",
    })


# ---------------------------------------------------------------
# AI Grade Predictions
# ---------------------------------------------------------------

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def student_grade_prediction(request, student_id):
    """
    Predict a student's future grades per subject using linear regression.
    Returns trend and predicted percentage for each subject.
    """
    if request.user.role not in ('admin', 'teacher', 'parent', 'student'):
        return Response({'error': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)

    school = request.user.school
    try:
        student = Student.objects.get(id=student_id, user__school=school)
    except Student.DoesNotExist:
        return Response({'error': 'Student not found.'}, status=status.HTTP_404_NOT_FOUND)

    from .ml_predictions import predict_student_grades
    predictions = predict_student_grades(student)
    return Response({'predictions': predictions, 'student': student.user.full_name})


# ── Timetable Conflict Detection ──────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def timetable_conflict_check(request):
    """
    Check the school's timetable for conflicts.
    Returns a list of detected conflicts:
      - Teacher double-booked (same teacher, same day, overlapping slots)
      - Room double-booked (same room, same day, overlapping slots)
      - Class double-booked (same class, same day, overlapping slots)
    """
    if request.user.role not in ('admin', 'superadmin'):
        return Response({'error': 'Admins only'}, status=status.HTTP_403_FORBIDDEN)

    school = request.user.school
    entries = list(
        Timetable.objects.filter(class_assigned__school=school)
        .select_related('class_assigned', 'subject', 'teacher__user')
        .order_by('day_of_week', 'start_time')
    )

    def overlaps(a, b):
        """True if two timetable entries overlap in time."""
        return a.start_time < b.end_time and b.start_time < a.end_time

    conflicts = []

    # Group by day for efficiency
    from itertools import combinations
    from collections import defaultdict

    by_day = defaultdict(list)
    for e in entries:
        by_day[e.day_of_week].append(e)

    for day, day_entries in by_day.items():
        for a, b in combinations(day_entries, 2):
            if not overlaps(a, b):
                continue

            # Teacher conflict
            if a.teacher_id and a.teacher_id == b.teacher_id:
                conflicts.append({
                    'type': 'teacher',
                    'day': day,
                    'teacher': f"{a.teacher.user.first_name} {a.teacher.user.last_name}",
                    'slot_1': {'class': a.class_assigned.name, 'subject': a.subject.name,
                               'time': f"{a.start_time}-{a.end_time}"},
                    'slot_2': {'class': b.class_assigned.name, 'subject': b.subject.name,
                               'time': f"{b.start_time}-{b.end_time}"},
                })

            # Room conflict (skip empty rooms)
            if a.room and b.room and a.room.strip().lower() == b.room.strip().lower():
                conflicts.append({
                    'type': 'room',
                    'day': day,
                    'room': a.room,
                    'slot_1': {'class': a.class_assigned.name, 'subject': a.subject.name,
                               'time': f"{a.start_time}-{a.end_time}"},
                    'slot_2': {'class': b.class_assigned.name, 'subject': b.subject.name,
                               'time': f"{b.start_time}-{b.end_time}"},
                })

            # Class double-booked
            if a.class_assigned_id == b.class_assigned_id:
                conflicts.append({
                    'type': 'class',
                    'day': day,
                    'class': a.class_assigned.name,
                    'slot_1': {'subject': a.subject.name, 'time': f"{a.start_time}-{a.end_time}"},
                    'slot_2': {'subject': b.subject.name, 'time': f"{b.start_time}-{b.end_time}"},
                })

    return Response({
        'total_entries': len(entries),
        'conflict_count': len(conflicts),
        'conflicts': conflicts,
    })


# ── Subject-Teacher Assignment ─────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def subject_teachers(request, subject_id):
    """Get all teachers assigned to a subject"""
    if request.user.role not in ('admin', 'hr', 'superadmin'):
        return Response({'error': 'Admin only'}, status=status.HTTP_403_FORBIDDEN)
    try:
        subject = Subject.objects.get(id=subject_id, school=request.user.school)
    except Subject.DoesNotExist:
        return Response({'error': 'Subject not found'}, status=status.HTTP_404_NOT_FOUND)

    teachers = subject.teachers.select_related('user').all()
    data = [{
        'id': t.id,
        'user_id': t.user.id,
        'first_name': t.user.first_name,
        'last_name': t.user.last_name,
        'email': t.user.email,
        'qualification': t.qualification,
    } for t in teachers]
    return Response(data)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def assign_teacher_to_subject(request, subject_id):
    """Assign a teacher to a subject"""
    if request.user.role not in ('admin', 'hr', 'superadmin'):
        return Response({'error': 'Admin only'}, status=status.HTTP_403_FORBIDDEN)
    try:
        subject = Subject.objects.get(id=subject_id, school=request.user.school)
    except Subject.DoesNotExist:
        return Response({'error': 'Subject not found'}, status=status.HTTP_404_NOT_FOUND)

    teacher_id = request.data.get('teacher_id')
    if not teacher_id:
        return Response({'error': 'teacher_id is required'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        from .models import Teacher
        teacher = Teacher.objects.get(id=teacher_id, user__school=request.user.school)
    except Teacher.DoesNotExist:
        return Response({'error': 'Teacher not found'}, status=status.HTTP_404_NOT_FOUND)

    if subject.teachers.filter(id=teacher.id).exists():
        return Response({'error': 'Teacher already assigned to this subject'}, status=status.HTTP_400_BAD_REQUEST)

    teacher.subjects_taught.add(subject)
    return Response({'message': f'{teacher.user.first_name} {teacher.user.last_name} assigned to {subject.name}'}, status=status.HTTP_201_CREATED)


@api_view(['DELETE'])
@permission_classes([permissions.IsAuthenticated])
def remove_teacher_from_subject(request, subject_id, teacher_id):
    """Remove a teacher from a subject"""
    if request.user.role not in ('admin', 'hr', 'superadmin'):
        return Response({'error': 'Admin only'}, status=status.HTTP_403_FORBIDDEN)
    try:
        subject = Subject.objects.get(id=subject_id, school=request.user.school)
    except Subject.DoesNotExist:
        return Response({'error': 'Subject not found'}, status=status.HTTP_404_NOT_FOUND)

    try:
        from .models import Teacher
        teacher = Teacher.objects.get(id=teacher_id)
    except Teacher.DoesNotExist:
        return Response({'error': 'Teacher not found'}, status=status.HTTP_404_NOT_FOUND)

    teacher.subjects_taught.remove(subject)
    return Response({'message': f'{teacher.user.first_name} {teacher.user.last_name} removed from {subject.name}'})


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def subject_class_assignments(request, subject_id):
    """List class assignments for a subject within the user's school."""
    if request.user.role not in ('admin', 'hr', 'superadmin'):
        return Response({'error': 'Admin only'}, status=status.HTTP_403_FORBIDDEN)
    school = request.user.school
    subject = Subject.objects.filter(id=subject_id, school=school).first()
    if not subject:
        return Response({'error': 'Subject not found'}, status=status.HTTP_404_NOT_FOUND)

    qs = ClassSubjectAssignment.objects.filter(
        school=school, subject=subject
    ).select_related('class_obj', 'teacher__user').order_by('class_obj__name', 'academic_year')
    data = [{
        "id": a.id,
        "class_id": a.class_obj_id,
        "class_name": a.class_obj.name,
        "academic_year": a.academic_year,
        "is_core": a.is_core,
        "teacher_id": a.teacher_id,
        "teacher_name": a.teacher.user.get_full_name() if a.teacher and a.teacher.user else None,
    } for a in qs]
    return Response(data)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def assign_subject_to_classes(request, subject_id):
    """Assign one subject to one or many classes."""
    if request.user.role not in ('admin', 'hr', 'superadmin'):
        return Response({'error': 'Admin only'}, status=status.HTTP_403_FORBIDDEN)
    school = request.user.school
    subject = Subject.objects.filter(id=subject_id, school=school).first()
    if not subject:
        return Response({'error': 'Subject not found'}, status=status.HTTP_404_NOT_FOUND)

    class_ids = request.data.get('class_ids') or []
    if isinstance(class_ids, str):
        class_ids = [x.strip() for x in class_ids.split(',') if x.strip()]
    if not isinstance(class_ids, list) or not class_ids:
        return Response({'error': 'class_ids is required (non-empty list).'}, status=status.HTTP_400_BAD_REQUEST)

    teacher = None
    teacher_id = request.data.get('teacher_id')
    if teacher_id not in (None, '', 'null'):
        teacher = Teacher.objects.filter(id=teacher_id, user__school=school).first()
        if not teacher:
            return Response({'error': 'Teacher not found in this school.'}, status=status.HTTP_400_BAD_REQUEST)

    academic_year = (request.data.get('academic_year') or '').strip()
    is_core = _parse_bool(request.data.get('is_core'))
    duplicate_strategy = (request.data.get('duplicate_strategy') or 'skip').strip().lower()
    if duplicate_strategy not in ('skip', 'update', 'error'):
        duplicate_strategy = 'skip'

    created, updated = 0, 0
    errors = []
    assignments = []

    for raw_id in class_ids:
        try:
            class_id = int(raw_id)
        except (TypeError, ValueError):
            errors.append({"class_id": raw_id, "error": "Invalid class id."})
            continue

        class_obj = Class.objects.filter(id=class_id, school=school).first()
        if not class_obj:
            errors.append({"class_id": class_id, "error": "Class not found in this school."})
            continue

        row_year = academic_year or class_obj.academic_year or str(timezone.now().year)
        existing = ClassSubjectAssignment.objects.filter(
            school=school, class_obj=class_obj, subject=subject, academic_year=row_year
        ).first()
        if existing:
            if duplicate_strategy == 'skip':
                continue
            if duplicate_strategy == 'error':
                errors.append({"class_id": class_id, "error": f"Assignment already exists for {class_obj.name} ({row_year})."})
                continue
            existing.teacher = teacher
            existing.is_core = is_core
            existing.save()
            updated += 1
            assignments.append(existing.id)
        else:
            a = ClassSubjectAssignment.objects.create(
                school=school,
                class_obj=class_obj,
                subject=subject,
                teacher=teacher,
                academic_year=row_year,
                is_core=is_core,
                created_by=request.user,
            )
            created += 1
            assignments.append(a.id)

    return Response({
        "subject_id": subject.id,
        "created": created,
        "updated": updated,
        "errors": errors,
        "assignment_ids": assignments,
        "message": f"Assigned {subject.code} to classes. Created {created}, updated {updated}, errors {len(errors)}."
    })


@api_view(['DELETE'])
@permission_classes([permissions.IsAuthenticated])
def remove_subject_class_assignment(request, subject_id, assignment_id):
    if request.user.role not in ('admin', 'hr', 'superadmin'):
        return Response({'error': 'Admin only'}, status=status.HTTP_403_FORBIDDEN)
    school = request.user.school
    subject = Subject.objects.filter(id=subject_id, school=school).first()
    if not subject:
        return Response({'error': 'Subject not found'}, status=status.HTTP_404_NOT_FOUND)
    assignment = ClassSubjectAssignment.objects.filter(
        id=assignment_id, school=school, subject=subject
    ).first()
    if not assignment:
        return Response({'error': 'Assignment not found'}, status=status.HTTP_404_NOT_FOUND)
    assignment.delete()
    return Response({'message': 'Class assignment removed.'})


# ── Report Card Publishing ─────────────────────────────────────────────────

def _announce_report_release(author, class_obj, term, year):
    """Notify class users that report cards are now available."""
    from users.models import Notification
    from .models import ParentChildLink

    # Announcement feed entries
    for audience in ['student', 'parent', 'teacher']:
        Announcement.objects.create(
            title=f'Report Cards Available — {term} {year}',
            content=(
                f'Report cards for {class_obj.name} ({term} {year}) are now available for download. '
                f'Go to your Results page to download the PDF.'
            ),
            author=author,
            target_audience=audience,
            target_audiences=[audience],
            target_class=class_obj,
        )

    student_users = [s.user for s in Student.objects.filter(student_class=class_obj).select_related('user')]
    parent_users = [
        link.parent.user
        for link in ParentChildLink.objects.filter(
            student__student_class=class_obj,
            is_confirmed=True,
        ).select_related('parent__user')
    ]

    unique_recipients = {u.id: u for u in [*student_users, *parent_users] if u}
    payload = [
        Notification(
            user=u,
            title=f'Report Cards Available — {term} {year}',
            message=f'Your report card for {class_obj.name} is now available.',
            notification_type='general',
            link='/student/results' if u.role == 'student' else '/parent/performance',
        )
        for u in unique_recipients.values()
    ]
    if payload:
        Notification.objects.bulk_create(payload)


def _excluded_student_ids_for_delivery(school, class_obj, year, term):
    from .models import ReportCardDeliveryExclusion
    return set(
        ReportCardDeliveryExclusion.objects.filter(
            school=school,
            class_obj=class_obj,
            academic_year=year,
            academic_term=term,
        ).values_list('student_id', flat=True)
    )


def _publish_class_reports(school, class_obj, year, term, actor, access_scope='all'):
    from .models import ReportCardRelease
    release, created = ReportCardRelease.objects.get_or_create(
        school=school,
        class_obj=class_obj,
        academic_year=year,
        academic_term=term,
        defaults={'published_by': actor, 'access_scope': access_scope},
    )
    if not created and release.access_scope != access_scope:
        release.access_scope = access_scope
        release.save(update_fields=['access_scope'])
    if created:
        _announce_report_release(actor, class_obj, term, year)
    excluded_count = len(_excluded_student_ids_for_delivery(school, class_obj, year, term))
    return release, created, excluded_count


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def generate_reports_for_teachers(request):
    """
    Step 1 workflow gate:
    Admin generates class/year/term report batch before teachers can submit sign-off.
    Body: { "class_id": 5, "year": "2026", "term": "Term 1" }
    """
    if request.user.role != 'admin':
        return Response({'error': 'Only admins can generate report batches'}, status=status.HTTP_403_FORBIDDEN)

    class_id = request.data.get('class_id')
    year = _normalize_report_year(request.data.get('year'))
    term = _normalize_report_term(request.data.get('term'))
    if not all([class_id, year, term]):
        return Response({'error': 'class_id, year, and term are required'}, status=status.HTTP_400_BAD_REQUEST)

    school = request.user.school
    try:
        class_obj = Class.objects.get(id=class_id, school=school)
    except Class.DoesNotExist:
        return Response({'error': 'Class not found'}, status=status.HTTP_404_NOT_FOUND)

    from .models import ReportCardGeneration
    generation, created = ReportCardGeneration.objects.get_or_create(
        school=school,
        class_obj=class_obj,
        academic_year=year,
        academic_term=term,
        defaults={'generated_by': request.user},
    )
    if not created:
        generation.generated_by = request.user
        generation.generated_at = timezone.now()
        generation.save(update_fields=['generated_by', 'generated_at'])
        return Response({
            'message': f'Report batch regenerated for {class_obj.name} - {term} {year}.',
            'generated': True,
            'already_generated': True,
        }, status=status.HTTP_200_OK)

    return Response({
        'message': f'Report batch generated for teachers: {class_obj.name} - {term} {year}.',
        'generated': True,
        'already_generated': False,
    }, status=status.HTTP_201_CREATED)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def publish_reports(request):
    """
    Publish report cards for a single class/year/term.
    Creates a ReportCardRelease record and sends announcements to
    students, parents, and the class teacher.
    Body: { "class_id": 5, "year": "2026", "term": "Term 1" }
    """
    if request.user.role != 'admin':
        return Response({'error': 'Only admins can publish reports'}, status=status.HTTP_403_FORBIDDEN)

    class_id = request.data.get('class_id')
    year = _normalize_report_year(request.data.get('year'))
    term = _normalize_report_term(request.data.get('term'))

    access_scope = (request.data.get('access_scope') or 'all').strip().lower()
    if access_scope not in ('all', 'fully_paid'):
        return Response({'error': "access_scope must be 'all' or 'fully_paid'."}, status=status.HTTP_400_BAD_REQUEST)

    if not all([class_id, year, term]):
        return Response({'error': 'class_id, year, and term are required'}, status=status.HTTP_400_BAD_REQUEST)

    school = request.user.school
    try:
        class_obj = Class.objects.get(id=class_id, school=school)
    except Class.DoesNotExist:
        return Response({'error': 'Class not found'}, status=status.HTTP_404_NOT_FOUND)

    from .models import ReportCardGeneration
    if not ReportCardGeneration.objects.filter(
        school=school,
        class_obj=class_obj,
        academic_year=year,
        academic_term=term,
    ).exists():
        return Response(
            {'error': 'Generate this class report batch first before publishing.'},
            status=status.HTTP_400_BAD_REQUEST
        )

    from .models import ReportCardApprovalRequest
    approval = ReportCardApprovalRequest.objects.filter(
        school=school,
        class_obj=class_obj,
        academic_year=year,
        academic_term=term,
    ).first()
    if not approval:
        return Response(
            {'error': 'No teacher sign-off submission exists for this class/year/term yet.'},
            status=status.HTTP_400_BAD_REQUEST,
        )
    if approval.status == 'rejected':
        return Response(
            {'error': 'This report batch was rejected. Ask the teacher to resubmit first.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    release, created, excluded_count = _publish_class_reports(
        school, class_obj, year, term, request.user, access_scope=access_scope
    )
    if not created:
        return Response({'message': f'Reports for {class_obj.name} - {term} {year} were already published',
                         'already_published': True})

    approval.status = 'approved'
    approval.reviewed_by = request.user
    approval.reviewed_at = timezone.now()
    approval.admin_note = (request.data.get('admin_note') or '').strip()
    approval.save(update_fields=['status', 'reviewed_by', 'reviewed_at', 'admin_note'])
    if approval.requested_by:
        from users.models import Notification
        Notification.objects.create(
            user=approval.requested_by,
            title='Report Sign-off Approved',
            message=f"{class_obj.name} ({term} {year}) has been approved and published.",
            notification_type='general',
            link='/teacher/results',
        )

    return Response({
        'message': f'Reports published for {class_obj.name} - {term} {year}',
        'class_name': class_obj.name,
        'access_scope': release.access_scope,
        'excluded_students_count': excluded_count,
        'published': True,
    }, status=status.HTTP_201_CREATED)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def publish_all_reports(request):
    """
    Publish report cards for ALL classes in the school for a given year/term.
    Body: { "year": "2026", "term": "Term 1" }
    """
    if request.user.role != 'admin':
        return Response({'error': 'Only admins can publish reports'}, status=status.HTTP_403_FORBIDDEN)

    year = _normalize_report_year(request.data.get('year'))
    term = _normalize_report_term(request.data.get('term'))

    access_scope = (request.data.get('access_scope') or 'all').strip().lower()
    if access_scope not in ('all', 'fully_paid'):
        return Response({'error': "access_scope must be 'all' or 'fully_paid'."}, status=status.HTTP_400_BAD_REQUEST)

    if not all([year, term]):
        return Response({'error': 'year and term are required'}, status=status.HTTP_400_BAD_REQUEST)

    school = request.user.school
    classes = Class.objects.filter(school=school)

    from .models import ReportCardApprovalRequest, ReportCardGeneration
    published = []
    published_details = []
    skipped = []

    for class_obj in classes:
        is_generated = ReportCardGeneration.objects.filter(
            school=school,
            class_obj=class_obj,
            academic_year=year,
            academic_term=term,
        ).exists()
        if not is_generated:
            skipped.append(class_obj.name)
            continue
        approval = ReportCardApprovalRequest.objects.filter(
            school=school,
            class_obj=class_obj,
            academic_year=year,
            academic_term=term,
            status='pending',
        ).first()
        if not approval:
            skipped.append(class_obj.name)
            continue
        _, created, excluded_count = _publish_class_reports(
            school,
            class_obj,
            year,
            term,
            request.user,
            access_scope=access_scope,
        )
        if created:
            approval.status = 'approved'
            approval.reviewed_by = request.user
            approval.reviewed_at = timezone.now()
            approval.save(update_fields=['status', 'reviewed_by', 'reviewed_at'])
            label = class_obj.name
            if excluded_count:
                label = f"{label} ({excluded_count} excluded)"
            published.append(label)
            published_details.append({
                'class_id': class_obj.id,
                'class_name': class_obj.name,
                'excluded_students_count': excluded_count,
            })
        else:
            skipped.append(class_obj.name)

    return Response({
        'message': f'{len(published)} class(es) published, {len(skipped)} already published',
        'published_classes': published,
        'published_details': published_details,
        'skipped_classes': skipped,
    }, status=status.HTTP_201_CREATED)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def list_published_reports(request):
    """List all published report card releases for the school."""
    if request.user.role != 'admin':
        return Response({'error': 'Only admins can view this'}, status=status.HTTP_403_FORBIDDEN)

    from .models import ReportCardRelease
    releases = ReportCardRelease.objects.filter(
        school=request.user.school
    ).select_related('class_obj', 'published_by').order_by('-published_at')

    data = [{
        'id': r.id,
        'class_id': r.class_obj.id,
        'class_name': r.class_obj.name,
        'academic_year': r.academic_year,
        'academic_term': r.academic_term,
        'access_scope': r.access_scope,
        'published_by': r.published_by.full_name,
        'published_at': r.published_at.isoformat(),
    } for r in releases]

    return Response({'releases': data})


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def list_report_approval_requests(request):
    """Admin list of report sign-off requests by class/year/term."""
    if request.user.role not in ('admin', 'superadmin'):
        return Response({'error': 'Only admins can view this'}, status=status.HTTP_403_FORBIDDEN)

    from .models import ReportCardApprovalRequest, ReportCardDeliveryExclusion
    qs = ReportCardApprovalRequest.objects.filter(
        school=request.user.school
    ).select_related('class_obj', 'requested_by', 'reviewed_by').order_by('-submitted_at')

    status_filter = request.query_params.get('status')
    year = _normalize_report_year(request.query_params.get('year'))
    term = _normalize_report_term(request.query_params.get('term'))
    class_id = request.query_params.get('class_id')
    if status_filter:
        qs = qs.filter(status=status_filter)
    if year:
        qs = qs.filter(academic_year=year)
    if term:
        qs = qs.filter(academic_term=term)
    if class_id:
        qs = qs.filter(class_obj_id=class_id)

    data = []
    for r in qs:
        students = Student.objects.filter(student_class_id=r.class_obj_id).select_related('user')
        exclusion_ids = set(
            ReportCardDeliveryExclusion.objects.filter(
                school=request.user.school,
                class_obj_id=r.class_obj_id,
                academic_year=r.academic_year,
                academic_term=r.academic_term,
            ).values_list('student_id', flat=True)
        )
        data.append({
            'id': r.id,
            'class_id': r.class_obj_id,
            'class_name': r.class_obj.name,
            'academic_year': r.academic_year,
            'academic_term': r.academic_term,
            'status': r.status,
            'submitted_at': r.submitted_at.isoformat(),
            'requested_by': r.requested_by.full_name if r.requested_by else None,
            'reviewed_at': r.reviewed_at.isoformat() if r.reviewed_at else None,
            'reviewed_by': r.reviewed_by.full_name if r.reviewed_by else None,
            'admin_note': r.admin_note,
            'excluded_students_count': len(exclusion_ids),
            'students': [
                {
                    'id': s.id,
                    'full_name': s.user.full_name,
                    'student_number': s.user.student_number or '',
                    'excluded_data_issue': s.id in exclusion_ids,
                }
                for s in students
            ],
        })
    return Response({'requests': data})


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def set_report_delivery_exclusion(request):
    """Admin toggles 'data issue' exclusion for student report delivery."""
    if request.user.role not in ('admin', 'superadmin'):
        return Response({'error': 'Only admins can do this'}, status=status.HTTP_403_FORBIDDEN)

    class_id = request.data.get('class_id')
    student_id = request.data.get('student_id')
    year = _normalize_report_year(request.data.get('year'))
    term = _normalize_report_term(request.data.get('term'))
    excluded_raw = request.data.get('excluded', True)
    excluded = str(excluded_raw).strip().lower() in ('1', 'true', 'yes', 'on')

    if not all([class_id, student_id, year, term]):
        return Response({'error': 'class_id, student_id, year and term are required'}, status=status.HTTP_400_BAD_REQUEST)

    school = request.user.school
    try:
        class_obj = Class.objects.get(id=class_id, school=school)
        student = Student.objects.select_related('user').get(id=student_id, student_class=class_obj)
    except (Class.DoesNotExist, Student.DoesNotExist):
        return Response({'error': 'Class or student not found for this school'}, status=status.HTTP_404_NOT_FOUND)

    from .models import ReportCardDeliveryExclusion
    if excluded:
        _, created = ReportCardDeliveryExclusion.objects.get_or_create(
            school=school,
            class_obj=class_obj,
            student=student,
            academic_year=year,
            academic_term=term,
            defaults={'created_by': request.user, 'reason': ReportCardDeliveryExclusion.REASON_DATA_ISSUE},
        )
        action_word = 'added' if created else 'kept'
    else:
        ReportCardDeliveryExclusion.objects.filter(
            school=school,
            class_obj=class_obj,
            student=student,
            academic_year=year,
            academic_term=term,
        ).delete()
        action_word = 'removed'

    log_school_audit(
        user=request.user,
        action='UPDATE',
        model_name='ReportCardDeliveryExclusion',
        object_id=student.id,
        object_repr=f"{student.user.full_name} / {class_obj.name} / {term} {year}",
        changes={'excluded': excluded, 'reason': 'data_issue', 'action': action_word},
        status_code=200,
        ip_address=request.META.get('REMOTE_ADDR'),
    )

    return Response({'message': f'Delivery exclusion {action_word}.', 'excluded': excluded}, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def review_report_approval_request(request, request_id):
    """Admin approves or rejects a teacher report submission."""
    if request.user.role not in ('admin', 'superadmin'):
        return Response({'error': 'Only admins can review this'}, status=status.HTTP_403_FORBIDDEN)

    from .models import ReportCardApprovalRequest
    try:
        approval = ReportCardApprovalRequest.objects.select_related('class_obj').get(
            id=request_id, school=request.user.school
        )
    except ReportCardApprovalRequest.DoesNotExist:
        return Response({'error': 'Request not found'}, status=status.HTTP_404_NOT_FOUND)

    decision = (request.data.get('decision') or '').strip().lower()
    admin_note = (request.data.get('admin_note') or '').strip()
    if decision not in ('approve', 'reject'):
        return Response({'error': "decision must be 'approve' or 'reject'"}, status=status.HTTP_400_BAD_REQUEST)

    if decision == 'reject':
        approval.status = 'rejected'
        approval.reviewed_by = request.user
        approval.reviewed_at = timezone.now()
        approval.admin_note = admin_note
        approval.save(update_fields=['status', 'reviewed_by', 'reviewed_at', 'admin_note'])
        if approval.requested_by:
            from users.models import Notification
            Notification.objects.create(
                user=approval.requested_by,
                title='Report Sent Back For Correction',
                message=(
                    f"{approval.class_obj.name} ({approval.academic_term} {approval.academic_year}) "
                    f"was sent back by admin. {admin_note or 'Please update feedback and resubmit.'}"
                ),
                notification_type='general',
                link='/teacher/report-feedback',
            )
        return Response({'message': 'Report request sent back to teacher for corrections.', 'status': approval.status})

    access_scope = (request.data.get('access_scope') or 'all').strip().lower()
    if access_scope not in ('all', 'fully_paid'):
        return Response({'error': "access_scope must be 'all' or 'fully_paid'."}, status=status.HTTP_400_BAD_REQUEST)

    from .models import ReportCardGeneration
    if not ReportCardGeneration.objects.filter(
        school=request.user.school,
        class_obj=approval.class_obj,
        academic_year=approval.academic_year,
        academic_term=approval.academic_term,
    ).exists():
        return Response(
            {'error': 'Generate this class report batch first before final sign-off.'},
            status=status.HTTP_400_BAD_REQUEST
        )

    _, created, excluded_count = _publish_class_reports(
        request.user.school,
        approval.class_obj,
        approval.academic_year,
        approval.academic_term,
        request.user,
        access_scope=access_scope,
    )
    approval.status = 'approved'
    approval.reviewed_by = request.user
    approval.reviewed_at = timezone.now()
    approval.admin_note = admin_note
    approval.save(update_fields=['status', 'reviewed_by', 'reviewed_at', 'admin_note'])
    if approval.requested_by:
        from users.models import Notification
        Notification.objects.create(
            user=approval.requested_by,
            title='Report Sign-off Approved',
            message=(
                f"{approval.class_obj.name} ({approval.academic_term} {approval.academic_year}) "
                "has been approved and published."
            ),
            notification_type='general',
            link='/teacher/results',
        )

    return Response({
        'message': 'Report request approved and reports are now visible to parents/students.',
        'status': approval.status,
        'published_now': created,
        'excluded_students_count': excluded_count,
    })


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def admin_at_risk_students(request):
    """
    Admin view: All at-risk students with comprehensive filtering and details.
    
    Query params:
        view: 'overall' (default) or 'by_subject'
        subject_id: Filter by specific subject (only with view=by_subject)
        search: Search by name, email, or student number
        risk_level: 'all' (default), 'high', 'medium', 'low'
        class_id: Filter by class
        sort_by: 'risk_score' (default), 'name', 'date'
    """
    if request.user.role not in ('admin', 'hr', 'superadmin', 'teacher'):
        return Response({'error': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
    
    school = request.user.school
    view_type = request.query_params.get('view', 'overall')
    search = request.query_params.get('search', '').strip()
    subject_id = request.query_params.get('subject_id', None)
    class_id = request.query_params.get('class_id', None)
    sort_by = request.query_params.get('sort_by', 'risk_score')
    
    # Get students
    students = Student.objects.filter(user__school=school, user__is_active=True).select_related('user', 'student_class')
    if request.user.role == 'teacher':
        teacher_profile = Teacher.objects.filter(user=request.user).first()
        taught_class_ids = set(
            Class.objects.filter(class_teacher=request.user, school=school).values_list('id', flat=True)
        )
        if teacher_profile:
            taught_class_ids.update(
                teacher_profile.teaching_classes.filter(school=school).values_list('id', flat=True)
            )
        if not taught_class_ids:
            return Response({
                'students': [],
                'top_performers': [],
                'total_at_risk': 0,
                'view_type': view_type,
                'filter': {
                    'search': search,
                    'subject_id': subject_id,
                    'class_id': class_id,
                }
            })
        students = students.filter(student_class_id__in=list(taught_class_ids))
    
    if class_id:
        try:
            students = students.filter(student_class_id=int(class_id))
        except (ValueError, TypeError):
            pass
    
    # Search filter
    if search:
        students = students.filter(
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(user__email__icontains=search) |
            Q(user__student_number__icontains=search)
        )
    
    at_risk_data = []
    non_risk_data = []
    from .ml_predictions import predict_student_grades
    from .at_risk_alerts import get_student_risk_score
    
    for student in students:
        try:
            predictions = predict_student_grades(student)
            
            # Calculate overall risk score
            overall_risk_score = get_student_risk_score(student)
            at_risk_subjects = [p for p in predictions if p['at_risk']]
            total_subjects = len(predictions)
            
            if view_type == 'by_subject' and subject_id:
                # Filter to specific subject
                subject_preds = [p for p in predictions if p['subject_id'] == int(subject_id)]
                if not subject_preds or not subject_preds[0]['at_risk']:
                    continue
                
                for pred in subject_preds:
                    # Get recent alerts
                    from .models import AtRiskAlert
                    recent_alerts = list(
                        AtRiskAlert.objects.filter(
                            student=student,
                            subject_id=int(subject_id)
                        ).order_by('-created_at')[:3].values(
                            'id', 'status', 'created_at', 'triggered_by', 'current_grade', 'predicted_grade'
                        )
                    )
                    
                    entry = {
                        'student_id': student.id,
                        'name': student.user.full_name,
                        'student_number': student.user.student_number or '',
                        'email': student.user.email,
                        'class': student.student_class.name,
                        'subject': pred['subject'],
                        'current_grade': pred['current_grade'],
                        'current_percentage': round(pred['current_percentage'], 1),
                        'predicted_grade': pred['predicted_grade'],
                        'predicted_percentage': round(pred['predicted_percentage'], 1),
                        'trend': pred['trend'],
                        'confidence': pred['confidence'],
                        'intervention': pred['intervention'],
                        'overall_risk_score': overall_risk_score,
                        'recent_alerts': recent_alerts,
                    }
                    at_risk_data.append(entry)
            else:
                # Overall view - show all at-risk subjects for students if they have any
                if at_risk_subjects:
                    from .models import AtRiskAlert
                    recent_alerts = list(
                        AtRiskAlert.objects.filter(
                            student=student,
                            status__in=['new', 'acknowledged', 'intervention_scheduled']
                        ).order_by('-created_at')[:3].values(
                            'id', 'subject__name', 'status', 'created_at', 'triggered_by', 'current_grade'
                        )
                    )
                    
                    entry = {
                        'student_id': student.id,
                        'name': student.user.full_name,
                        'student_number': student.user.student_number or '',
                        'email': student.user.email,
                        'class': student.student_class.name,
                        'overall_risk_score': overall_risk_score,
                        'at_risk_count': len(at_risk_subjects),
                        'total_subjects': total_subjects,
                        'at_risk_subjects': [
                            {
                                'subject': p['subject'],
                                'current_grade': p['current_grade'],
                                'predicted_grade': p['predicted_grade'],
                                'trend': p['trend'],
                                'percentage': round(p['current_percentage'], 1),
                            }
                            for p in at_risk_subjects
                        ],
                        'recent_alerts': recent_alerts,
                    }
                    at_risk_data.append(entry)
                else:
                    avg_pct = 0.0
                    if predictions:
                        avg_pct = round(
                            sum(float(p.get('current_percentage', 0) or 0) for p in predictions) / len(predictions), 1
                        )
                    non_risk_data.append({
                        'student_id': student.id,
                        'name': student.user.full_name,
                        'student_number': student.user.student_number or '',
                        'email': student.user.email,
                        'class': student.student_class.name,
                        'overall_risk_score': overall_risk_score,
                        'average_percentage': avg_pct,
                        'total_subjects': total_subjects,
                    })
        
        except Exception as e:
            logger.error(f"Error processing student {student.id} for at-risk view: {str(e)}")
            continue
    
    # Sort
    if sort_by == 'name':
        at_risk_data.sort(key=lambda x: x['name'])
    elif sort_by == 'date':
        # Sort by most recently created alerts
        at_risk_data.sort(
            key=lambda x: x['recent_alerts'][0]['created_at'] if x['recent_alerts'] else timezone.now(),
            reverse=True
        )
    else:  # risk_score (default)
        at_risk_data.sort(key=lambda x: x['overall_risk_score'], reverse=True)
    
    non_risk_data.sort(
        key=lambda x: (x.get('average_percentage', 0), -x.get('overall_risk_score', 0)),
        reverse=True
    )
    top_performers = non_risk_data[:10]

    return Response({
        'students': at_risk_data,
        'top_performers': top_performers,
        'total_at_risk': len(at_risk_data),
        'view_type': view_type,
        'filter': {
            'search': search,
            'subject_id': subject_id,
            'class_id': class_id,
        }
    })
    if check_rate_limit(request, group='admin_parent_link_approve', rate='30/m'):
        return Response({'error': 'Too many requests. Please try again shortly.'}, status=status.HTTP_429_TOO_MANY_REQUESTS)
    if check_rate_limit(request, group='admin_parent_link_decline', rate='30/m'):
        return Response({'error': 'Too many requests. Please try again shortly.'}, status=status.HTTP_429_TOO_MANY_REQUESTS)


@api_view(['GET', 'POST'])
@permission_classes([permissions.IsAuthenticated])
def attendance_permissions(request):
    """List or create attendance permission entries (admin/HR only)."""
    if request.user.role not in ('admin', 'hr', 'superadmin'):
        return Response({'error': 'Only admin/HR can manage attendance permissions.'},
                        status=status.HTTP_403_FORBIDDEN)

    school = request.user.school
    if request.method == 'GET':
        student_id = request.query_params.get('student_id')
        date_val = request.query_params.get('date')
        qs = (AttendancePermission.objects
              .filter(class_assigned__school=school)
              .select_related('student__user', 'class_assigned', 'approved_by', 'created_by')
              .order_by('-date', '-id'))
        if student_id:
            qs = qs.filter(student_id=student_id)
        if date_val:
            qs = qs.filter(date=date_val)
        return Response([
            {
                'id': p.id,
                'student_id': p.student_id,
                'student_name': p.student.user.full_name,
                'class_id': p.class_assigned_id,
                'class_name': p.class_assigned.name,
                'date': p.date.isoformat(),
                'period_number': p.period_number,
                'period_label': p.period_label,
                'reason': p.reason,
                'approved': p.approved,
                'approved_by': p.approved_by.full_name if p.approved_by else None,
                'created_by': p.created_by.full_name if p.created_by else None,
            }
            for p in qs[:200]
        ])

    student_id = request.data.get('student_id')
    class_id = request.data.get('class_id')
    date_val = request.data.get('date')
    period_number = request.data.get('period_number')
    period_label = (request.data.get('period_label') or '').strip()
    reason = (request.data.get('reason') or '').strip()
    approved = bool(request.data.get('approved', True))

    if not student_id or not class_id or not date_val:
        return Response({'error': 'student_id, class_id, and date are required.'},
                        status=status.HTTP_400_BAD_REQUEST)
    try:
        from datetime import datetime
        parsed_date = datetime.strptime(date_val, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return Response({'error': 'Invalid date format. Use YYYY-MM-DD.'},
                        status=status.HTTP_400_BAD_REQUEST)
    try:
        student = Student.objects.get(id=student_id, user__school=school)
        class_obj = Class.objects.get(id=class_id, school=school)
    except (Student.DoesNotExist, Class.DoesNotExist):
        return Response({'error': 'Student or class not found.'}, status=status.HTTP_404_NOT_FOUND)

    parsed_period_number = None
    if period_number not in (None, ''):
        try:
            parsed_period_number = int(period_number)
            if parsed_period_number < 1:
                raise ValueError
        except (TypeError, ValueError):
            return Response({'error': 'period_number must be a positive integer.'},
                            status=status.HTTP_400_BAD_REQUEST)

    permission = AttendancePermission.objects.create(
        student=student,
        class_assigned=class_obj,
        date=parsed_date,
        period_number=parsed_period_number,
        period_label=period_label,
        reason=reason,
        approved=approved,
        approved_by=request.user if approved else None,
        created_by=request.user,
    )
    return Response({'id': permission.id, 'message': 'Attendance permission saved.'}, status=status.HTTP_201_CREATED)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def attendance_period_tracking_start_date(request):
    """Set effective date from which period-level bunk detection is enforced."""
    if request.user.role not in ('admin', 'hr', 'superadmin'):
        return Response({'error': 'Only admin/HR can update this setting.'},
                        status=status.HTTP_403_FORBIDDEN)
    raw_date = request.data.get('start_date')
    if not raw_date:
        return Response({'error': 'start_date is required (YYYY-MM-DD).'},
                        status=status.HTTP_400_BAD_REQUEST)
    try:
        from datetime import datetime
        parsed_date = datetime.strptime(raw_date, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return Response({'error': 'Invalid date format. Use YYYY-MM-DD.'},
                        status=status.HTTP_400_BAD_REQUEST)

    settings, _ = SchoolSettings.objects.get_or_create(school=request.user.school)
    settings.attendance_period_tracking_start_date = parsed_date
    settings.save(update_fields=['attendance_period_tracking_start_date'])
    return Response({
        'message': 'Period attendance tracking start date updated.',
        'start_date': settings.attendance_period_tracking_start_date.isoformat(),
    })


def _attendance_admin_allowed(user):
    return user.role in ('admin', 'hr', 'superadmin')


@api_view(['PATCH'])
@permission_classes([permissions.IsAuthenticated])
def edit_class_attendance(request, attendance_id):
    """Admin/HR edit for daily class attendance records."""
    if not _attendance_admin_allowed(request.user):
        return Response({'error': 'Only admin/HR can edit class attendance.'},
                        status=status.HTTP_403_FORBIDDEN)
    record = ClassAttendance.objects.filter(
        id=attendance_id,
        class_assigned__school=request.user.school,
    ).select_related('student__user', 'class_assigned').first()
    if not record:
        return Response({'error': 'Attendance record not found.'}, status=status.HTTP_404_NOT_FOUND)

    status_value = request.data.get('status')
    remarks = request.data.get('remarks')
    if status_value is not None:
        if status_value not in {'present', 'absent', 'late', 'excused'}:
            return Response({'error': 'Invalid status value.'}, status=status.HTTP_400_BAD_REQUEST)
        record.status = status_value
    if remarks is not None:
        record.remarks = str(remarks)
    record.recorded_by = request.user
    record.save(update_fields=['status', 'remarks', 'recorded_by'])

    return Response({
        'message': 'Class attendance updated.',
        'id': record.id,
        'student_id': record.student_id,
        'date': record.date.isoformat(),
        'status': record.status,
        'remarks': record.remarks or '',
    })


@api_view(['PATCH'])
@permission_classes([permissions.IsAuthenticated])
def edit_subject_attendance(request, attendance_id):
    """Admin/HR edit for subject/period attendance records."""
    if not _attendance_admin_allowed(request.user):
        return Response({'error': 'Only admin/HR can edit subject attendance.'},
                        status=status.HTTP_403_FORBIDDEN)
    record = SubjectAttendance.objects.filter(
        id=attendance_id,
        class_assigned__school=request.user.school,
    ).select_related('student__user', 'class_assigned').first()
    if not record:
        return Response({'error': 'Attendance record not found.'}, status=status.HTTP_404_NOT_FOUND)

    status_value = request.data.get('status')
    remarks = request.data.get('remarks')
    if status_value is not None:
        if status_value not in {'present', 'absent', 'late', 'excused'}:
            return Response({'error': 'Invalid status value.'}, status=status.HTTP_400_BAD_REQUEST)
        record.status = status_value
    if remarks is not None:
        record.remarks = str(remarks)

    has_permission = AttendancePermission.objects.filter(
        student=record.student,
        class_assigned=record.class_assigned,
        date=record.date,
        approved=True,
    ).filter(
        Q(period_number=record.period_number) |
        Q(period_number__isnull=True)
    ).exists()
    settings = SchoolSettings.objects.filter(school=request.user.school).first()
    tracking_active = bool(
        settings and settings.attendance_period_tracking_start_date and
        record.date >= settings.attendance_period_tracking_start_date
    )
    daily = ClassAttendance.objects.filter(student=record.student, date=record.date).first()
    bunk_flag = bool(
        tracking_active and
        record.status == 'absent' and
        not has_permission and
        daily and daily.status in ('present', 'late')
    )

    record.marked_with_permission = has_permission
    record.bunk_flag = bunk_flag
    record.bunk_reason = 'Absent during period without approved permission' if bunk_flag else ''
    record.recorded_by = request.user
    record.save(update_fields=[
        'status', 'remarks', 'marked_with_permission', 'bunk_flag', 'bunk_reason', 'recorded_by',
    ])

    return Response({
        'message': 'Subject attendance updated.',
        'id': record.id,
        'student_id': record.student_id,
        'date': record.date.isoformat(),
        'status': record.status,
        'remarks': record.remarks or '',
        'period_number': record.period_number,
        'period_label': record.period_label,
        'marked_with_permission': record.marked_with_permission,
        'bunk_flag': record.bunk_flag,
        'bunk_reason': record.bunk_reason,
    })
